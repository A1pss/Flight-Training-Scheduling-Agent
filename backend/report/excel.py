"""四表 Excel 渲染（v6 §10.1~§10.4）。

## 版式契约只有一份，在校验器那边

工作表名与顺序、三张表的表头、Sheet 4 的七个区块标题与列名、机组角色后缀、
语义开关的序列化分隔符 —— 全部 **import 自 `backend.validator.workbook`**，
本模块一个字都不另抄。理由很实在：那边是**回读**（闸门3），这边是**写出**，
两边各留一份常量迟早漂，而漂了以后表现为「深度相等断言过不了」，
排查成本远高于一行 import。

```
excel.py  写出  ──→  xlsx  ──→  workbook.py 回读 ──→ SchedulePlan
   ↑                                                      │
   └────────────── deep_diff 必须为空（§4.3 闸门3）────────┘
```

## 版式基准的采信边界（§10.5，业务方 2026-08-13 确认）

`docs/M3_版式基准抽取清单.md` 是本模块的设计依据。落到代码里的三条：

1. **只采信版式，内容一律不采信** —— `image 1~4.png` 的架次数据在本模块里
   一个字节都没有，全部来自 `ReportBundle`。
2. **业务方裁决：不着色**。类别底纹与 A 类红字下划线都不做，正文一律黑字白底；
   **唯一保留的底纹**是 Sheet 4 区块标题行的浅色底 —— 那是 v6 §10.4 的强制项，
   不是版式图带来的。
3. **Sheet 1~3 不加跑道列**（§10.4 区块7 的原始动机）：`runway_id` 与
   `is_recurrent` 只在 Sheet 4 区块 7 出现，否则就偏离版式基准了。

## 时间列为什么必须是文本

`HH:MM` 一旦写成 `datetime.time`，openpyxl 会给它一个日期格式，在别的软件里
显示成 `0.25` 这种 Excel 序列号。v6 §4.3 点名要断言这件事，
`workbook._require_time_text` 会当场把它判成格式错误。所以本模块所有时间列
都写字符串并显式设 `number_format="@"`（文本）。
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date as date_type
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from backend.core.ruleset import IDENTITY_STUDENT
from backend.report.bundle import ReportBundle
from backend.schemas.plan import CrewMember, Sortie
from backend.schemas.validation import RULE_IDS
from backend.validator.context import AircraftFacts
from backend.validator.schema import WEEKDAY_ORDER
from backend.validator.workbook import (
    BLOCK2_HEADERS,
    BLOCK3_HEADERS,
    BLOCK4_HEADERS,
    BLOCK7_HEADERS,
    BLOCK_TITLES,
    META_COVER,
    META_ISO_WEEK,
    META_PLAN_ID,
    META_RULESET,
    META_RUNWAY_MODEL,
    META_SEMANTICS,
    META_SHA,
    META_SNAPSHOT,
    META_SWITCHES,
    META_TIER,
    NOT_RECURRENT_MARK,
    RECURRENT_MARK,
    ROLE_SUFFIX,
    SHEET1_HEADERS,
    SHEET2_HEADERS,
    SHEET3_HEADERS,
    SHEET_ORDER,
    SWITCH_KV,
    SWITCH_SEP,
)

# ─────────────────────────────────────────────────────────────────────
# 版式参数（`docs/M3_版式基准抽取清单.md` §2，业务方 2026-08-13 确认）
# ─────────────────────────────────────────────────────────────────────
#: 正文字体：image 4 是四张图里唯一的真 Excel 截图，字体形态为宋体
BODY_FONT = "宋体"
#: 等宽列（Sheet 2 时间列 / Sheet 3 起飞列，§10.2「等宽字体」）
MONO_FONT = "Consolas"
FONT_SIZE = 11

#: Sheet 4 区块标题行的浅色底纹 —— v6 §10.4 的强制项，与版式图无关
BLOCK_TITLE_FILL = "DDEBF6"

#: 列宽（image 4 实测列宽比 1:1:1:2.9:1，换算成 Excel 字符宽）
SHEET1_WIDTHS = (12, 12, 12, 35, 16)
SHEET2_WIDTHS = (14, 8, 16, 32, 16)
SHEET3_WIDTHS = (10, 8, 10, 32, 20)
SHEET4_WIDTHS = (22, 46, 14, 12, 12, 40, 14, 14, 14, 10)

#: Sheet 4 最宽的区块是区块3（10 列）
SHEET4_COLUMNS = len(BLOCK3_HEADERS)

#: 机组列的角色排序：教员在前，受训人在后（版式基准 `孙军教，陈伟学`）
_ROLE_ORDER: Mapping[str, int] = {"教员": 0, "学员": 1, "单飞": 1, "复训": 1}

#: 空值一律显示为破折号（`上次执行` 为 `—` 表示锚点 NULL，按 S-12 从本周周一起算）
DASH = "—"

#: 一周天数（排班周恒为周一~周日）
WEEK_DAYS = 7

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ─────────────────────────────────────────────────────────────────────
# 拼接格式（与 `workbook.py` 的解析器互为逆运算）
# ─────────────────────────────────────────────────────────────────────
def fmt_time(value: time) -> str:
    """`HH:MM`。秒被丢弃是刻意的——版式基准与契约都只到分钟。"""
    return f"{value.hour:02d}:{value.minute:02d}"


def fmt_mission(name: str, mission_id: str, airspace: str | None = None) -> str:
    """`本场起落航线 (missionA-1)（Small Area A）`；Sheet 2/3 不带空域段。"""
    head = f"{name} ({mission_id})"
    return f"{head}（{airspace}）" if airspace else head


def sorted_crew(crew: Sequence[CrewMember]) -> tuple[CrewMember, ...]:
    """教员在前、受训人在后；同序时按编号 —— 报告必须逐字节可复现（铁律 9）。"""
    return tuple(sorted(crew, key=lambda c: (_ROLE_ORDER.get(c.role, 9), c.person_id)))


def fmt_crew_sheet1(crew: Sequence[CrewMember]) -> str:
    """`孙军教，陈伟学` / `何超单` / `刘斌训`。

    分隔符是**全角逗号**（业务方 2026-08-13 裁定；v6 §10.1 正文原写「顿号」，
    与同节示例、版式基准、回读解析器三者不符，已同步改正）。
    """
    return "，".join(f"{c.name}{ROLE_SUFFIX[c.role]}" for c in sorted_crew(crew))


def fmt_crew_sheet3(crew: Sequence[CrewMember]) -> str:
    """`(高超/罗磊)`；单飞 `(何超)`；复训 `(刘斌)`（只写姓名，不写角色）。"""
    return "(" + "/".join(c.name for c in sorted_crew(crew)) + ")"


def fmt_aircraft_role(aircraft_id: str, role: str) -> str:
    """`(AC49/学员)`、`(AC10/单飞)`、`(AC84/复训)`。"""
    return f"({aircraft_id}/{role})"


def fmt_switches(switches: Mapping[str, str]) -> str:
    """`S-01=all_missions_completed；S-02=class_level；…`（序列化口径由回读侧固化）。"""
    return SWITCH_SEP.join(f"{k}{SWITCH_KV}{switches[k]}" for k in sorted(switches))


def fmt_person(person_id: str, name: str) -> str:
    """Sheet 4 的人员列写作 `何超(P08)` —— Sheet 4 无版式基准，故直接带编号。"""
    return f"{name}({person_id})"


def minutes_between(start: time, end: time) -> int:
    return (end.hour * 60 + end.minute) - (start.hour * 60 + start.minute)


def minimum_required(freq_days: int, *, gap_days: int | None, start_offset: int = 0) -> int:
    """按频率滑窗推出的**本周最少应排次数**（v6 §3.5.3 的通式 D-4 + S-12）。

    这是**展示口径**，不是约束判定 —— 判定在 `validator/checks.py::check_c13`，
    本函数只回答「区块3 的『本周应排』该写几」。两者不会打架：校验器判的是
    「任意连续 `freq_days` 天窗口内 ≥1 次」，本函数算的是满足它所需的最少次数。

    - `gap_days` 为 None：锚点缺失，按 S-12 从本周周一起算，首次截止日 `freq_days − 1`；
    - `gap_days` 有值：首次截止日 `max(0, freq_days − gap)`（D-4 通式）；
    - `start_offset`：复训自 `recurrent_since` 起算，窗口整体后移。
    """
    if freq_days <= 0:
        raise ValueError(f"freq_days 必须为正，实际 {freq_days}")
    first = (freq_days - 1) if gap_days is None else max(0, freq_days - gap_days)
    day = start_offset + first
    count = 0
    while day < WEEK_DAYS:
        count += 1
        day += freq_days
    return count


# ─────────────────────────────────────────────────────────────────────
# 单元格写入
# ─────────────────────────────────────────────────────────────────────
@dataclass(frozen=True)
class CellStyle:
    bold: bool = False
    mono: bool = False
    border: bool = True
    align: str = "center"
    fill: str | None = None


_PLAIN = CellStyle(border=False, align="left")
_GROUP = CellStyle(bold=True, border=False, align="left")
HEADER_STYLE = CellStyle(bold=True)
_DATA = CellStyle()
_TIME = CellStyle(mono=True)
_LABEL = CellStyle(bold=False, align="left")
BLOCK_TITLE_STYLE = CellStyle(bold=True, border=False, align="left", fill=BLOCK_TITLE_FILL)


def put_cell(ws: Worksheet, row: int, col: int, value: Any, style: CellStyle) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name=MONO_FONT if style.mono else BODY_FONT, size=FONT_SIZE, bold=style.bold)
    cell.alignment = Alignment(horizontal=style.align, vertical="center")
    if style.border:
        cell.border = _BORDER
    if style.fill:
        cell.fill = PatternFill("solid", fgColor=style.fill)
    if isinstance(value, str):
        cell.number_format = "@"


def _write_row(
    ws: Worksheet,
    row: int,
    values: Sequence[Any],
    style: CellStyle,
    *,
    overrides: Mapping[int, CellStyle] | None = None,
) -> int:
    for i, value in enumerate(values):
        put_cell(ws, row, i + 1, value, (overrides or {}).get(i, style))
    return row + 1


def set_column_widths(ws: Worksheet, widths: Sequence[int]) -> None:
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


def _day_sorties(sorties: Iterable[Sortie]) -> dict[str, list[Sortie]]:
    """按星期分组，组内按起飞时刻升序（§10.1）。"""
    grouped: dict[str, list[Sortie]] = {}
    for s in sorties:
        grouped.setdefault(s.weekday, []).append(s)
    return {
        day: sorted(grouped[day], key=lambda s: (s.takeoff, s.sortie_id))
        for day in WEEKDAY_ORDER
        if day in grouped
    }


# ─────────────────────────────────────────────────────────────────────
# Sheet 1 · 分日飞行计划表（§10.1）
# ─────────────────────────────────────────────────────────────────────
def _render_sheet1(ws: Worksheet, bundle: ReportBundle) -> None:
    set_column_widths(ws, SHEET1_WIDTHS)
    row = 1
    for day, sorties in _day_sorties(bundle.plan.sorties).items():
        # 日期分组：跨 A~E 合并的 banner（版式基准 image 4；回读侧只认「A 列有值、其余为空」）
        _write_row(ws, row, [day, None, None, None, None], _GROUP)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(SHEET1_HEADERS))
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        row = _write_row(ws, row, list(SHEET1_HEADERS), HEADER_STYLE)
        for s in sorties:
            airspace = bundle.ctx.airspaces.get(s.airspace_id)
            row = _write_row(
                ws,
                row,
                [
                    fmt_time(s.takeoff),
                    fmt_time(s.landing),
                    s.aircraft_id,
                    fmt_mission(
                        s.mission_name,
                        s.mission_id,
                        airspace.name if airspace else s.airspace_id,
                    ),
                    fmt_crew_sheet1(s.crew),
                ],
                _DATA,
                # 起飞列加粗（v6 §10.1 + image 1；image 4 未加粗，规格压版式图）
                overrides={0: CellStyle(bold=True)},
            )
        row += 1  # 组间空一行


# ─────────────────────────────────────────────────────────────────────
# Sheet 2 · 飞行员训练时间表（§10.2）
# ─────────────────────────────────────────────────────────────────────
def _render_sheet2(ws: Worksheet, bundle: ReportBundle) -> None:
    set_column_widths(ws, SHEET2_WIDTHS)
    by_person: dict[str, list[tuple[Sortie, CrewMember]]] = {}
    for s in bundle.plan.sorties:
        for member in s.crew:
            by_person.setdefault(member.person_id, []).append((s, member))

    row = 1
    for person_id in sorted(by_person):
        person = bundle.ctx.persons.get(person_id)
        name = person.name if person else by_person[person_id][0][1].name
        row = _write_row(ws, row, [name, None, None, None, None], _GROUP)
        row = _write_row(ws, row, list(SHEET2_HEADERS), HEADER_STYLE)
        entries = by_person[person_id]
        for day in WEEKDAY_ORDER:
            same_day = sorted(
                (e for e in entries if e[0].weekday == day),
                key=lambda e: (e[0].takeoff, e[0].sortie_id),
            )
            if not same_day:
                continue
            row = _write_row(ws, row, [None, day, None, None, None], _GROUP)
            for s, member in same_day:
                row = _write_row(
                    ws,
                    row,
                    [
                        None,
                        None,
                        f"{fmt_time(s.takeoff)}-{fmt_time(s.landing)}",
                        fmt_mission(s.mission_name, s.mission_id),
                        fmt_aircraft_role(s.aircraft_id, member.role),
                    ],
                    _DATA,
                    overrides={2: _TIME, 3: CellStyle(bold=True)},
                )
        row += 1


# ─────────────────────────────────────────────────────────────────────
# Sheet 3 · 飞机排班表（§10.3）
# ─────────────────────────────────────────────────────────────────────
def _render_sheet3(ws: Worksheet, bundle: ReportBundle) -> None:
    set_column_widths(ws, SHEET3_WIDTHS)
    by_aircraft: dict[str, list[Sortie]] = {}
    for s in bundle.plan.sorties:
        by_aircraft.setdefault(s.aircraft_id, []).append(s)

    row = 1
    for aircraft_id in sorted(by_aircraft):
        row = _write_row(ws, row, [aircraft_id, None, None, None, None], _GROUP)
        row = _write_row(ws, row, list(SHEET3_HEADERS), HEADER_STYLE)
        for day in WEEKDAY_ORDER:
            same_day = sorted(
                (s for s in by_aircraft[aircraft_id] if s.weekday == day),
                key=lambda s: (s.takeoff, s.sortie_id),
            )
            if not same_day:
                continue
            row = _write_row(ws, row, [None, day, None, None, None], _GROUP)
            for s in same_day:
                row = _write_row(
                    ws,
                    row,
                    [
                        None,
                        None,
                        fmt_time(s.takeoff),
                        fmt_mission(s.mission_name, s.mission_id),
                        fmt_crew_sheet3(s.crew),
                    ],
                    _DATA,
                    overrides={2: _TIME},
                )
        row += 1


# ─────────────────────────────────────────────────────────────────────
# Sheet 4 · 合规与解释报告（§10.4，七区块）
# ─────────────────────────────────────────────────────────────────────
def _fmt_float(value: float | None, digits: int = 3) -> str:
    return DASH if value is None else f"{value:.{digits}f}"


def _runway_model_text(bundle: ReportBundle) -> str:
    """`dual_runway（RWY-1: JL-8/JL-9；RWY-2: JL-8）` —— 括号内由跑道事实生成。

    回读侧按 `split("（")[0]` 取枚举值，括号部分是给人看的。**不写死机型**。
    """
    detail = SWITCH_SEP.join(
        f"{rid}: {'/'.join(sorted(bundle.ctx.runways[rid].aircraft_types))}"
        for rid in sorted(bundle.ctx.runways)
    )
    return f"{bundle.plan.runway_model}（{detail}）" if detail else bundle.plan.runway_model


def _block1_rows(bundle: ReportBundle, generated_at: datetime) -> list[tuple[str, str]]:
    plan, stats = bundle.plan, bundle.stats
    tier = plan.relaxation_tier
    return [
        (META_PLAN_ID, plan.plan_id),
        (META_ISO_WEEK, plan.iso_week),
        (META_COVER, f"{plan.week_start.isoformat()} ~ {plan.week_end.isoformat()}"),
        (META_SNAPSHOT, plan.snapshot_id),
        (META_RULESET, plan.ruleset_version),
        (META_SEMANTICS, plan.semantics_version),
        (META_SWITCHES, fmt_switches(plan.semantics_switches)),
        (META_RUNWAY_MODEL, _runway_model_text(bundle)),
        (META_TIER, f"Tier {tier}（全硬约束）" if tier == 0 else f"Tier {tier}"),
        ("求解状态", stats.status),
        ("求解耗时(s)", f"{stats.wall_time_ms / 1000.0:.2f}"),
        ("目标值", _fmt_float(stats.objective_value)),
        ("gap", _fmt_float(stats.gap, 6)),
        ("worker 数", str(stats.num_workers)),
        ("seed", str(stats.random_seed)),
        ("生成时间", generated_at.isoformat()),
        (META_SHA, plan.content_sha256),
    ]


def _block2_rows(bundle: ReportBundle, *, readback: str) -> list[list[str]]:
    """14 行规则 + 末行格式校验三层。"""
    by_rule = {r.rule_id: r for r in bundle.validation.results}
    rows: list[list[str]] = []
    for n, rule_id in enumerate(RULE_IDS, start=1):
        result = by_rule.get(rule_id)
        if result is None:
            # 没跑就是没跑（铁律 6）——不写「通过」，写「未校验」
            rows.append([f"约束{n}", rule_id, "⚠ 未校验", "0", "0", "本次运行未提供该条结果"])
            continue
        hard = [v for v in result.violations if v.severity == "HARD"]
        detail = "；".join(v.detail for v in result.violations[:3]) or DASH
        notes = "；".join(result.notes)
        rows.append(
            [
                f"约束{n}",
                result.rule_title,
                "✅ 通过" if result.passed and not hard else "❌ 违规",
                str(result.checked_items),
                str(len(hard)),
                f"{detail}｜{notes}" if notes else detail,
            ]
        )
    fmt = bundle.format_report
    schema_mark = DASH if fmt is None else ("✅" if not fmt.schema_errors else "❌")
    integrity_mark = (
        DASH
        if fmt is None
        else ("✅" if not (fmt.integrity_errors or fmt.cross_table_errors) else "❌")
    )
    rows.append(
        [
            "格式校验",
            f"Schema 层 {schema_mark} / 业务完整性层 {integrity_mark} / 产物回读层 {readback}",
            "✅ 通过" if readback == "✅" else "⚠ 见说明",
            DASH,
            DASH,
            "三层口径见 v6 §4.3",
        ]
    )
    return rows


@dataclass(frozen=True)
class _ProgressRow:
    person_id: str
    mission_ids: tuple[str, ...]
    label: str
    status: str
    freq_days: int
    required: int
    scheduled: int
    last_done: str
    cycle_left: str
    tier: str = DASH
    debt: int | None = None


def _status_text(status: str, *, recurrent: bool, prereq_met: bool) -> str:
    if recurrent:
        return "到期复训"
    if status == "COMPLETED":
        return "已完成"
    if not prereq_met:
        return "先修未满足"
    return "未完成"


def _progress_rows(bundle: ReportBundle) -> list[_ProgressRow]:
    """区块3 的行集合。

    两条构造规则：

    1. **普通行按 (人, 课目)**；**复训行按 (人, 类别) 合并** —— S-11 的粒度是类别
       （Z-8），逐门列会把「本周应排」算成两次，与校验器的判定口径对不上。
    2. `plan.debts` 里的每一条都必须有**恰好一行**承载，且只有这些行的「松弛档」
       非 `—`（回读侧据此反解 `TrainingDebt`，多一行少一行都会破坏深度相等）。
    """
    plan, ctx = bundle.plan, bundle.ctx
    scheduled: dict[tuple[str, str], int] = {}
    for s in plan.sorties:
        for member in s.crew:
            key = (member.person_id, s.mission_id)
            scheduled[key] = scheduled.get(key, 0) + 1

    rows: list[_ProgressRow] = []
    recurrent_groups: dict[tuple[str, str], list[str]] = {}
    for (person_id, mission_id), prog in sorted(ctx.progress.items()):
        mission = ctx.missions.get(mission_id)
        if mission is None:
            continue
        if prog.is_recurrent:
            recurrent_groups.setdefault((person_id, mission.mission_class), []).append(mission_id)
            continue
        if prog.completed:
            continue  # 已完成的课目不受约束13（S-03），不占版面
        gap = None if prog.last_done_date is None else (ctx.week_start - prog.last_done_date).days
        rows.append(
            _ProgressRow(
                person_id=person_id,
                mission_ids=(mission_id,),
                label=mission_id,
                status=_status_text(prog.status, recurrent=False, prereq_met=prog.prereq_met),
                freq_days=mission.freq_days,
                # 先修未满足的组合按约束13 被排除在外（§3.6 BLOCKED ≠ INFEASIBLE），
                # 本周对它**没有**要求 —— 写成「应排 1 / 欠账 1」会让排班员以为
                # 这是个 Tier 0 都没排上的欠账，而它其实在区块4 里有完整交代
                required=0
                if not prog.prereq_met
                else minimum_required(mission.freq_days, gap_days=gap),
                scheduled=scheduled.get((person_id, mission_id), 0),
                last_done=prog.last_done_date.isoformat() if prog.last_done_date else DASH,
                cycle_left=_cycle_left(prog.cycle_start, mission.cycle_weeks, ctx.week_start),
            )
        )

    for (person_id, mission_class), mission_ids in sorted(recurrent_groups.items()):
        ids = tuple(sorted(mission_ids))
        first = ctx.progress[(person_id, ids[0])]
        mission = ctx.missions[ids[0]]
        since = first.recurrent_since
        offset = max(0, (since - ctx.week_start).days) if since else 0
        anchors = [
            ctx.progress[(person_id, mid)].last_done_date
            for mid in ids
            if ctx.progress[(person_id, mid)].last_done_date is not None
        ]
        anchor = max((a for a in anchors if a is not None), default=None)
        gap = None if anchor is None else (ctx.week_start - anchor).days
        rows.append(
            _ProgressRow(
                person_id=person_id,
                mission_ids=ids,
                label=f"{' / '.join(ids)}（{mission_class} 类复训）",
                status="到期复训",
                freq_days=mission.freq_days,
                required=minimum_required(mission.freq_days, gap_days=gap, start_offset=offset),
                scheduled=sum(scheduled.get((person_id, mid), 0) for mid in ids),
                last_done=anchor.isoformat() if anchor else DASH,
                cycle_left=DASH,
            )
        )

    return _attach_debts(bundle, rows)


def _attach_debts(bundle: ReportBundle, rows: list[_ProgressRow]) -> list[_ProgressRow]:
    """把 `plan.debts` 挂到对应行上；挂不上的单独补一行。"""
    by_key = {
        (r.person_id, r.mission_ids[0]): i for i, r in enumerate(rows) if len(r.mission_ids) == 1
    }
    out = list(rows)
    for debt in sorted(bundle.plan.debts, key=lambda d: (d.person_id, d.mission_id)):
        idx = by_key.get((debt.person_id, debt.mission_id))
        mission = bundle.ctx.missions.get(debt.mission_id)
        if idx is None:
            out.append(
                _ProgressRow(
                    person_id=debt.person_id,
                    mission_ids=(debt.mission_id,),
                    label=debt.mission_id,
                    status="欠账",
                    freq_days=mission.freq_days if mission else 0,
                    required=debt.required,
                    scheduled=debt.scheduled,
                    last_done=DASH,
                    cycle_left=DASH,
                    tier=debt.relaxed_by,
                    debt=debt.debt,
                )
            )
            continue
        base = out[idx]
        out[idx] = _ProgressRow(
            person_id=base.person_id,
            mission_ids=base.mission_ids,
            label=base.label,
            status=base.status,
            freq_days=base.freq_days,
            required=debt.required,
            scheduled=debt.scheduled,
            last_done=base.last_done,
            cycle_left=base.cycle_left,
            tier=debt.relaxed_by,
            debt=debt.debt,
        )
    return out


def _cycle_left(cycle_start: date_type | None, cycle_weeks: int, week_start: date_type) -> str:
    if cycle_start is None or cycle_weeks <= 0:
        return DASH
    elapsed = (week_start - cycle_start).days // 7
    return str(max(0, cycle_weeks - elapsed))


def _block3_rows(bundle: ReportBundle) -> list[list[str]]:
    out: list[list[str]] = []
    for r in _progress_rows(bundle):
        person = bundle.ctx.persons.get(r.person_id)
        debt = r.debt if r.debt is not None else max(0, r.required - r.scheduled)
        out.append(
            [
                fmt_person(r.person_id, person.name if person else r.person_id),
                r.label,
                r.status,
                f"每 {r.freq_days} 天 ≥1" if r.freq_days else DASH,
                str(r.required),
                str(r.scheduled),
                str(debt),
                r.last_done,
                r.cycle_left,
                r.tier,
            ]
        )
    return out


def _block4_rows(bundle: ReportBundle) -> list[list[str]]:
    """阻塞项 —— 披露率 100% 是 v6 §0.3 的可测断言，一条都不许省。"""
    out: list[list[str]] = []
    for item in sorted(bundle.plan.blocked_items, key=lambda b: (b.person_id, b.mission_id)):
        person = bundle.ctx.persons.get(item.person_id)
        missing = "、".join(item.missing_prereqs) if item.missing_prereqs else DASH
        unlock = f"完成 {missing} 后" if item.missing_prereqs else DASH
        out.append(
            [
                fmt_person(item.person_id, person.name if person else item.person_id),
                item.mission_id,
                item.reason,
                missing,
                unlock,
            ]
        )
    return out


BLOCK5_HEADERS: tuple[str, ...] = ("对象", "架次", "飞行时长(分)", "利用率", "备注")


def _maintenance_minutes(aircraft: AircraftFacts, week_start: date_type) -> int:
    """本周内维护时段与每日可用窗的**相交分钟数**。

    用于把维护时间从利用率的分母里扣掉 —— AC73 周五全天定检，那一天它就不是
    「闲着」，把它算进可用时长会让利用率显得比实际低。
    """
    total = 0
    for window in aircraft.maintenance:
        for offset in range(WEEK_DAYS):
            day = week_start + timedelta(days=offset)
            start = datetime.combine(day, aircraft.daily_window_start)
            end = datetime.combine(day, aircraft.daily_window_end)
            overlap = min(end, window.end) - max(start, window.start)
            total += max(0, int(overlap.total_seconds() // 60))
    return total


def _block5_rows(bundle: ReportBundle) -> list[list[str]]:
    """飞机 / 人员（非学员）/ 空域 / 跑道逐行。

    ⚠️ 这里的「利用率」是**展示统计**，分母写在备注或本函数文档里，不参与任何判定：

    - 飞机：飞行时长 / (7 天 × 每日可用窗 − 维护时段与可用窗的相交部分)；
    - 人员：架次 / 每周架次上限（约束11，按身份取值）；
    - 空域：占用时长 / (容量 × 7 天 × 该空域绑定课目的可用窗)；
    - 跑道：起降次数与 20 分钟窗口峰值占用（容量见约束9）。
    """
    ctx, plan = bundle.ctx, bundle.plan
    rows: list[list[str]] = []

    for aircraft_id in sorted(ctx.aircraft):
        ac = ctx.aircraft[aircraft_id]
        flights = [s for s in plan.sorties if s.aircraft_id == aircraft_id]
        minutes = sum(minutes_between(s.takeoff, s.landing) for s in flights)
        window = minutes_between(ac.daily_window_start, ac.daily_window_end)
        available = window * WEEK_DAYS - _maintenance_minutes(ac, plan.week_start)
        maintenance = "；".join(
            f"{w.start.date().isoformat()} {'全天' if w.all_day else ''}{w.kind}"
            for w in ac.maintenance
        )
        rows.append(
            [
                f"{aircraft_id}（{ac.aircraft_type}）",
                str(len(flights)),
                str(minutes),
                _pct(minutes, available),
                maintenance or DASH,
            ]
        )

    for person in ctx.sorted_persons():
        if person.identity == IDENTITY_STUDENT:
            continue  # 学员的进度在区块3，此处只列教员与成熟飞行员（v6 §10.4 区块5）
        flights = [s for s in plan.sorties if any(c.person_id == person.person_id for c in s.crew)]
        minutes = sum(minutes_between(s.takeoff, s.landing) for s in flights)
        cap = ctx.ruleset.weekly_sortie_cap(person.identity)
        notes: list[str] = []
        if person.unavailable_dates:
            notes.append(
                "不可用：" + "、".join(d.isoformat() for d in sorted(person.unavailable_dates))
            )
        for cls in sorted(person.qualifications):
            qual = person.qualifications[cls]
            if qual.expiry_date is not None:
                notes.append(f"{cls} 类 {qual.expiry_date.isoformat()} 到期")
        rows.append(
            [
                f"{person.name}({person.person_id})／{person.identity}",
                str(len(flights)),
                str(minutes),
                f"{len(flights)}/{cap}",
                "；".join(notes) or DASH,
            ]
        )

    day_window = minutes_between(ctx.ruleset.window_start, ctx.ruleset.window_end)
    for airspace_id in sorted(ctx.airspaces):
        asp = ctx.airspaces[airspace_id]
        flights = [s for s in plan.sorties if s.airspace_id == airspace_id]
        minutes = sum(minutes_between(s.takeoff, s.landing) for s in flights)
        available = asp.capacity * WEEK_DAYS * day_window
        rows.append(
            [
                f"{airspace_id}（{asp.name}）",
                str(len(flights)),
                str(minutes),
                _pct(minutes, available),
                f"同时段容量 {asp.capacity}，硬约束，见约束6",
            ]
        )

    window_min = ctx.ruleset.density_window_minutes
    cap = ctx.ruleset.density_window_cap
    for runway_id in sorted(ctx.runways):
        rwy = ctx.runways[runway_id]
        flights = [s for s in plan.sorties if s.runway_id == runway_id]
        rows.append(
            [
                f"{runway_id}（{rwy.name}）",
                f"{len(flights)} 起降",
                DASH,
                f"峰值 {_peak_window(flights, window_min)}/{cap}（{window_min} 分钟窗口）",
                "服务机型：" + "/".join(sorted(rwy.aircraft_types)),
            ]
        )
    return rows


def _pct(part: int, whole: int) -> str:
    return DASH if whole <= 0 else f"{part / whole * 100:.1f}%"


def _peak_window(flights: Sequence[Sortie], window_minutes: int) -> int:
    """同一跑道、同一天，`[t, t+window)` 内的最大起飞数（展示统计，非判定）。"""
    peak = 0
    by_day: dict[date_type, list[int]] = {}
    for s in flights:
        by_day.setdefault(s.date, []).append(s.takeoff.hour * 60 + s.takeoff.minute)
    for minutes in by_day.values():
        for anchor in minutes:
            peak = max(peak, sum(1 for m in minutes if anchor <= m < anchor + window_minutes))
    return peak


#: S-11 授权改写声明的兜底文本（校验器给出 notes 时优先用它给的）
S11_DECLARATION_LABEL = "授权改写声明"
S11_DECLARATION_FALLBACK = (
    "S-11：成熟飞行员到期资质转复训，系对 rules.pdf 约束2 字面语义的业务方授权改写"
    "（2026-08-06 裁定），非校验器漏判"
)


def _block6_rows(bundle: ReportBundle) -> list[tuple[str, str]]:
    relax = (
        "；".join(
            f"Tier {r.tier} {r.action}（代价：{r.cost}／{r.authority}）" for r in bundle.relaxations
        )
        or "本次未使用任何松弛"
    )
    notes = [n for n in bundle.validation.all_notes() if n.strip()]
    if bundle.s11_enabled:
        declaration = "；".join(notes) if notes else S11_DECLARATION_FALLBACK
    else:
        declaration = "S-11 开关为 off，本周期不适用"
    approval = bundle.approval
    return [
        ("使用的松弛", relax),
        ("冲突集", bundle.conflict_summary or "无"),
        (S11_DECLARATION_LABEL, declaration),
        ("审批人", approval.approver or DASH),
        ("审批时间", approval.approved_at.isoformat() if approval.approved_at else DASH),
    ]


def _block7_rows(bundle: ReportBundle) -> list[list[str]]:
    """`runway_id` 与 `is_recurrent` 的**唯一**落点（§10.4 / §4.3）。"""
    out: list[list[str]] = []
    for s in sorted(bundle.plan.sorties, key=lambda s: (s.date, s.takeoff, s.sortie_id)):
        ac = bundle.ctx.aircraft.get(s.aircraft_id)
        out.append(
            [
                s.sortie_id,
                s.date.isoformat(),
                fmt_time(s.takeoff),
                s.aircraft_id,
                ac.aircraft_type if ac else DASH,
                s.runway_id,
                s.airspace_id,
                RECURRENT_MARK if s.is_recurrent else NOT_RECURRENT_MARK,
            ]
        )
    return out


def _render_sheet4(ws: Worksheet, bundle: ReportBundle, *, readback: str) -> None:
    set_column_widths(ws, SHEET4_WIDTHS)
    row = 1

    def title(index: int) -> None:
        nonlocal row
        _write_row(ws, row, [BLOCK_TITLES[index]], BLOCK_TITLE_STYLE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SHEET4_COLUMNS)
        for col in range(1, SHEET4_COLUMNS + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=BLOCK_TITLE_FILL)
        row += 1

    def table(
        headers: Sequence[str], rows: Sequence[Sequence[str]], *, left: Sequence[int] = ()
    ) -> None:
        """一个区块的表格。`left` 是要左对齐的列（长文本列居中会很难读）。"""
        nonlocal row
        overrides = dict.fromkeys((0, *left), _LABEL)
        row = _write_row(ws, row, list(headers), HEADER_STYLE)
        for values in rows:
            row = _write_row(ws, row, list(values), _DATA, overrides=overrides)
        row += 1

    def pairs(items: Sequence[tuple[str, str]]) -> None:
        nonlocal row
        for label, value in items:
            row = _write_row(ws, row, [label, value], _PLAIN)
        row += 1

    title(0)
    pairs(_block1_rows(bundle, bundle.generated_at))
    title(1)
    table(BLOCK2_HEADERS, _block2_rows(bundle, readback=readback), left=(1, 5))
    title(2)
    table(BLOCK3_HEADERS, _block3_rows(bundle), left=(1,))
    title(3)
    table(BLOCK4_HEADERS, _block4_rows(bundle), left=(2, 3, 4))
    title(4)
    table(BLOCK5_HEADERS, _block5_rows(bundle), left=(4,))
    title(5)
    pairs(_block6_rows(bundle))
    title(6)
    table(BLOCK7_HEADERS, _block7_rows(bundle))


# ─────────────────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────────────────
def render_workbook(
    path: Path, bundle: ReportBundle, *, readback_passed: bool | None = None
) -> Path:
    """渲染四表并落盘。

    `readback_passed` 是**区块2 末行「产物回读层」那一格**的取值，三态：

    - `None` → 写「待回读」：文件刚写出、闸门3 还没跑，此时说「通过」就是
      报告一个未实际计算的结果（铁律 6）；
    - `True` / `False` → 写 ✅ / ❌，由 `report.verify.export_workbook` 在第一遍
      回读通过之后重渲一遍填进去（详见那边的两遍写出说明）。
    """
    wb = Workbook()
    default_sheet = wb.active  # 新建 Workbook 自带一张 "Sheet"，不要它
    if default_sheet is not None:
        wb.remove(default_sheet)
    sheets = [wb.create_sheet(name) for name in SHEET_ORDER]
    _render_sheet1(sheets[0], bundle)
    _render_sheet2(sheets[1], bundle)
    _render_sheet3(sheets[2], bundle)
    mark = {None: "待回读", True: "✅", False: "❌"}[readback_passed]
    _render_sheet4(sheets[3], bundle, readback=mark)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()
    return path


def sheet_titles() -> tuple[str, ...]:
    """模板与产物共用的工作表名（顺序固定，§10）。"""
    return SHEET_ORDER


__all__ = [
    "BLOCK5_HEADERS",
    "BLOCK_TITLE_STYLE",
    "BODY_FONT",
    "HEADER_STYLE",
    "MONO_FONT",
    "S11_DECLARATION_FALLBACK",
    "S11_DECLARATION_LABEL",
    "CellStyle",
    "fmt_aircraft_role",
    "fmt_crew_sheet1",
    "fmt_crew_sheet3",
    "fmt_mission",
    "fmt_person",
    "fmt_switches",
    "fmt_time",
    "minimum_required",
    "minutes_between",
    "put_cell",
    "render_workbook",
    "set_column_widths",
    "sheet_titles",
    "sorted_crew",
]
