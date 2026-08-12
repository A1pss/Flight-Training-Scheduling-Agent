"""Excel 模板与模板契约（v6 §10 / §10.7）。

## 「模板驱动」在本项目里是什么意思

四张表里 Sheet 1~3 的表头是**逐个分组重复**的（版式基准 image 1/4：每个星期、
每个人、每架飞机各带一行表头），所以模板不可能是「一张画好的空表，往里填行」。
本模块把模板拆成两件东西：

1. **模板契约** :func:`template_contract` —— 工作表名与顺序、三张表的表头、
   Sheet 4 的七个区块标题与各自的列名、列宽、字体。渲染层与回读层都按它走，
   契约测试拿它当断言基准。
2. **模板文件** `templates/plan_workbook_template.xlsx` —— 由 :func:`build_template`
   从同一份契约生成的空壳，四个工作表、顺序固定、表头就位。它的用途是
   **给业务方看版式**、以及让「产物的表头是否逐字匹配模板」这句话有个实体可比。

两者同源，所以不会漂：契约变了，模板重新生成一次即可（`python -m backend.report.template`）。

模板里**不放任何架次数据** —— `data/origin/image 1~4.png` 的内容一律不采信
（v6 §1.2.2 / §10.5，`docs/M3_版式基准抽取清单.md` 列了 17 条违反硬约束的证据）。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from backend.core.config import PROJECT_ROOT
from backend.report.excel import (
    BLOCK5_HEADERS,
    BLOCK_TITLE_STYLE,
    BODY_FONT,
    FONT_SIZE,
    HEADER_STYLE,
    MONO_FONT,
    SHEET1_WIDTHS,
    SHEET2_WIDTHS,
    SHEET3_WIDTHS,
    SHEET4_COLUMNS,
    SHEET4_WIDTHS,
    put_cell,
    set_column_widths,
)
from backend.validator.workbook import (
    BLOCK2_HEADERS,
    BLOCK3_HEADERS,
    BLOCK4_HEADERS,
    BLOCK7_HEADERS,
    BLOCK_TITLES,
    REQUIRED_META_LABELS,
    SHEET1_HEADERS,
    SHEET2_HEADERS,
    SHEET3_HEADERS,
    SHEET_ORDER,
)

#: 模板文件的交付位置（v6 §10.7：业务方确认后的版式基准清单也归档在这里）
TEMPLATE_PATH: Path = PROJECT_ROOT / "templates" / "plan_workbook_template.xlsx"

#: Sheet 4 里带表格的区块 → 其列名。区块1 与区块6 是「标签/取值」两列，无表头
BLOCK_TABLE_HEADERS: dict[str, tuple[str, ...]] = {
    BLOCK_TITLES[1]: BLOCK2_HEADERS,
    BLOCK_TITLES[2]: BLOCK3_HEADERS,
    BLOCK_TITLES[3]: BLOCK4_HEADERS,
    BLOCK_TITLES[4]: BLOCK5_HEADERS,
    BLOCK_TITLES[6]: BLOCK7_HEADERS,
}


def template_contract() -> dict[str, Any]:
    """模板契约的 JSON 形态（契约测试与前端都可以消费）。"""
    return {
        "sheets": list(SHEET_ORDER),
        "headers": {
            SHEET_ORDER[0]: list(SHEET1_HEADERS),
            SHEET_ORDER[1]: list(SHEET2_HEADERS),
            SHEET_ORDER[2]: list(SHEET3_HEADERS),
        },
        "blocks": list(BLOCK_TITLES),
        "block_headers": {title: list(cols) for title, cols in BLOCK_TABLE_HEADERS.items()},
        "meta_labels": list(REQUIRED_META_LABELS),
        "widths": {
            SHEET_ORDER[0]: list(SHEET1_WIDTHS),
            SHEET_ORDER[1]: list(SHEET2_WIDTHS),
            SHEET_ORDER[2]: list(SHEET3_WIDTHS),
            SHEET_ORDER[3]: list(SHEET4_WIDTHS),
        },
        "fonts": {"body": BODY_FONT, "mono": MONO_FONT, "size": FONT_SIZE},
    }


def build_template(path: Path | None = None) -> Path:
    """生成模板空壳。四个工作表、顺序固定、表头就位、**无任何数据行**。"""
    target = path or TEMPLATE_PATH
    wb = Workbook()
    default_sheet = wb.active  # 新建 Workbook 自带一张 "Sheet"，不要它
    if default_sheet is not None:
        wb.remove(default_sheet)
    sheets = [wb.create_sheet(name) for name in SHEET_ORDER]
    for ws, headers, widths in zip(
        sheets[:3],
        (SHEET1_HEADERS, SHEET2_HEADERS, SHEET3_HEADERS),
        (SHEET1_WIDTHS, SHEET2_WIDTHS, SHEET3_WIDTHS),
        strict=True,
    ):
        set_column_widths(ws, widths)
        for i, title in enumerate(headers):
            put_cell(ws, 1, i + 1, title, HEADER_STYLE)

    ws4 = sheets[3]
    set_column_widths(ws4, SHEET4_WIDTHS)
    row = 1
    for title in BLOCK_TITLES:
        put_cell(ws4, row, 1, title, BLOCK_TITLE_STYLE)
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SHEET4_COLUMNS)
        row += 1
        for i, col in enumerate(BLOCK_TABLE_HEADERS.get(title, ())):
            put_cell(ws4, row, i + 1, col, HEADER_STYLE)
        if title in BLOCK_TABLE_HEADERS:
            row += 1
        row += 1

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    wb.close()
    return target


def read_template_headers(path: Path | None = None) -> dict[str, list[str]]:
    """从模板文件读回表头 —— 「产物表头逐字匹配模板」这句话的实体依据。"""
    wb = load_workbook(path or TEMPLATE_PATH, data_only=True)
    out: dict[str, list[str]] = {}
    for name in SHEET_ORDER[:3]:
        ws = wb[name]
        out[name] = [str(c.value) for c in next(ws.iter_rows(max_row=1)) if c.value is not None]
    wb.close()
    return out


def main() -> None:  # pragma: no cover - CLI 入口
    path = build_template()
    print(f"模板已生成：{path}")


if __name__ == "__main__":  # pragma: no cover
    main()
