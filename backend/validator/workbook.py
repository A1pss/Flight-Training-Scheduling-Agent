"""格式校验第 3 层：Excel 产物回读与深度相等断言（v6 §4.3）。

```
闸门3  写出 xlsx → 回读 → 反解为 SchedulePlan → 与源对象 deep_diff 必须为空
```

这一层让格式通过率在**逻辑上**必然为 100%：写出的内容能被完整反解回原对象，
格式就不可能错。代价是 `Sortie` 的每一个字段都必须在四张表的某一处出现 ——
`runway_id` 与 `is_recurrent` 落在 **Sheet 4 区块 7**（v6 §10.4 / §4.3），
Sheet 1~3 不加跑道列以免偏离版式基准（§1.2.2）。

## 本模块定义的是「回读契约」，写出侧（M3/W5）必须照它写

Excel 写出在 M3 做。本模块把**反解所需的版式约定**固化成模块级常量与解析器：
工作表名与顺序、表头逐字内容、分组层次、单元格拼接格式、Sheet 4 各区块的标题与
列名。M3 直接 import 这些常量来写，两边就不会漂。

三处需要 M3 特别注意的约定（都是「反解不到就做不到深度相等」的字段）：

1. **区块 7 承载 `sortie_id` / `runway_id` / `airspace_id` / `is_recurrent`** ——
   Sheet 1~3 里没有它们（§10.4 的原始动机）。
2. **区块 1 增加一行「语义开关」** —— `SchedulePlan.semantics_switches` 参与
   `content_sha256`（v6 附录 B 脚注），不落表就反解不回来。§10.4 区块1 的字段
   列表给的是示例，Sheet 4 本身「无版式基准可依，由 §10 定义」（§10.5），
   故这里补一行是**扩充而非偏离**；与当初为承载 `runway_id` 新增区块 7 同一性质。
3. **Sheet 1~3 里人员只出现姓名**（版式基准如此）。`person_id` 由人员表
   （`ValidationContext`）按姓名反查；重名会被如实报成回读错误，不做猜测。

## 时间列必须是 `HH:MM` 文本

v6 §4.3 点名的断言：**时间列不得是 Excel 序列号**。openpyxl 读到 `datetime.time`
或 `float` 即判失败 —— 那种单元格在别的软件里会显示成 `0.25`，业务方在现场
一眼就能看出不对，但机器如果不查就会一路放过去。
"""

from __future__ import annotations

import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date as date_type
from datetime import datetime, timedelta
from datetime import time as time_type
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.schemas.plan import BlockedItem, CrewMember, SchedulePlan, Sortie, TrainingDebt
from backend.schemas.validation import SchemaCheckReport
from backend.validator.context import ValidationContext
from backend.validator.schema import (
    WEEKDAY_ORDER,
    SortieRow,
    ThreeTableProjection,
    check_cross_table_consistency,
)

# ─────────────────────────────────────────────────────────────────────
# 版式契约（M3 写出侧照此实现）
# ─────────────────────────────────────────────────────────────────────
#: 四张工作表，顺序固定（v6 §10）
SHEET_ORDER: tuple[str, ...] = (
    "分日飞行计划表",
    "飞行员训练时间表",
    "飞机排班表",
    "合规与解释报告",
)

SHEET1_HEADERS: tuple[str, ...] = ("起飞", "着陆", "飞机", "课目（空域）", "机组")
SHEET2_HEADERS: tuple[str, ...] = ("飞行员", "星期", "时间", "课目", "飞机/角色")
SHEET3_HEADERS: tuple[str, ...] = ("机号", "星期", "起飞", "课目", "机组")

#: Sheet 4 的七个区块标题（v6 §10.4）
BLOCK_TITLES: tuple[str, ...] = (
    "区块1 · 计划元信息",
    "区块2 · 约束校验结果",
    "区块3 · 训练进度与欠账",
    "区块4 · 阻塞项",
    "区块5 · 资源利用",
    "区块6 · 松弛与决策记录",
    "区块7 · 跑道与空域占用明细",
)
BLOCK2_HEADERS: tuple[str, ...] = ("规则编号", "规则名称", "判定", "检查项数", "违规数", "说明")
BLOCK3_HEADERS: tuple[str, ...] = (
    "人员",
    "课目",
    "状态",
    "频率要求",
    "本周应排",
    "实际排",
    "欠账",
    "上次执行",
    "周期剩余(周)",
    "松弛档",
)
BLOCK4_HEADERS: tuple[str, ...] = ("人员", "课目", "阻塞原因", "缺失先修", "预计解锁")
BLOCK7_HEADERS: tuple[str, ...] = (
    "架次号",
    "日期",
    "起飞",
    "机号",
    "机型",
    "跑道",
    "空域",
    "复训标记",
)

#: 区块1 反解所需的字段标签。缺一个就反解不回 `SchedulePlan`
META_PLAN_ID = "计划编号"
META_ISO_WEEK = "ISO 周"
META_COVER = "覆盖日期"
META_SNAPSHOT = "数据快照"
META_RULESET = "规则版本"
META_SEMANTICS = "语义版本"
META_SWITCHES = "语义开关"
META_RUNWAY_MODEL = "跑道模型"
META_TIER = "松弛档位"
META_SHA = "内容指纹"
REQUIRED_META_LABELS: tuple[str, ...] = (
    META_PLAN_ID,
    META_ISO_WEEK,
    META_COVER,
    META_SNAPSHOT,
    META_RULESET,
    META_SEMANTICS,
    META_SWITCHES,
    META_RUNWAY_MODEL,
    META_TIER,
    META_SHA,
)

#: 机组角色后缀（v6 §10.1：教员「教」、学员「学」、单飞「单」、复训「训」）
ROLE_SUFFIX: Mapping[str, str] = {"教员": "教", "学员": "学", "单飞": "单", "复训": "训"}
SUFFIX_ROLE: Mapping[str, str] = {v: k for k, v in ROLE_SUFFIX.items()}

#: 区块7「复训标记」列的取值
RECURRENT_MARK = "复训"
NOT_RECURRENT_MARK = "—"

#: 语义开关序列化：`S-01=all_missions_completed；S-02=class_level`
SWITCH_SEP = "；"
SWITCH_KV = "="

_TIME_RE = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")
_MISSION_RE = re.compile(r"^(?P<name>.+?)\s*\((?P<mid>mission[A-Z]-\d+)\)(（(?P<asp>.*)）)?$")
_TIER_RE = re.compile(r"(\d+)")
_RANGE_RE = re.compile(r"^\s*(\d{4}-\d{2}-\d{2})\s*~\s*(\d{4}-\d{2}-\d{2})\s*$")


# ─────────────────────────────────────────────────────────────────────
# 单元格取值与类型断言
# ─────────────────────────────────────────────────────────────────────
class WorkbookFormatError(Exception):
    """回读期发现的版式问题。由 `verify_workbook` 收集成 diff，不外抛。"""


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _require_time_text(value: Any, where: str) -> str:
    """时间列必须是 `HH:MM` **文本**，不是 Excel 序列号，也不是 `datetime.time`。"""
    if isinstance(value, time_type | datetime | float | int):
        raise WorkbookFormatError(
            f"{where} 的时间单元格类型为 {type(value).__name__}（{value!r}），"
            "必须写成 HH:MM 文本，否则在其他软件里会显示为 Excel 序列号"
        )
    text = _text(value)
    if not _TIME_RE.match(text):
        raise WorkbookFormatError(f"{where} 的时间格式非法：{text!r}（应为 HH:MM）")
    return text


def _parse_time(text: str) -> time_type:
    hh, mm = text.split(":")
    return time_type(int(hh), int(mm))


def _parse_date(value: Any, where: str) -> date_type:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_type):
        return value
    text = _text(value)
    try:
        return date_type.fromisoformat(text)
    except ValueError as exc:
        raise WorkbookFormatError(f"{where} 的日期格式非法：{text!r}（应为 YYYY-MM-DD）") from exc


def _row_values(ws: Worksheet, width: int) -> list[list[Any]]:
    return [list(row[:width]) for row in ws.iter_rows(values_only=True)]


def _is_blank(row: Sequence[Any]) -> bool:
    return all(_text(c) == "" for c in row)


# ─────────────────────────────────────────────────────────────────────
# 拼接格式的解析
# ─────────────────────────────────────────────────────────────────────
def parse_mission_cell(text: str, where: str) -> tuple[str, str, str | None]:
    """`本场起落航线 (missionA-1)（Small Area A）` → (课目名, 课目号, 空域名)。"""
    m = _MISSION_RE.match(text.strip())
    if not m:
        raise WorkbookFormatError(f"{where} 的课目单元格不合拼接格式：{text!r}")
    return m.group("name").strip(), m.group("mid"), m.group("asp")


def parse_sheet1_crew(text: str, where: str) -> tuple[tuple[str, str], ...]:
    """`孙军教，陈伟学` → ((姓名, 角色), ...)。"""
    out: list[tuple[str, str]] = []
    for token in [t for t in text.replace(",", "，").split("，") if t.strip()]:
        token = token.strip()
        role = SUFFIX_ROLE.get(token[-1])
        if role is None:
            raise WorkbookFormatError(
                f"{where} 的机组项 {token!r} 缺少角色后缀（合法后缀：{'/'.join(SUFFIX_ROLE)}）"
            )
        out.append((token[:-1], role))
    if not out:
        raise WorkbookFormatError(f"{where} 的机组列为空")
    return tuple(out)


def parse_paren_pair(text: str, where: str) -> tuple[str, ...]:
    """`(AC49/学员)` → ("AC49", "学员")；`(高超/罗磊)` → ("高超", "罗磊")。"""
    t = text.strip()
    if not (t.startswith("(") and t.endswith(")")):
        raise WorkbookFormatError(f"{where} 应为 (…/…) 形态，实际 {text!r}")
    return tuple(part.strip() for part in t[1:-1].split("/") if part.strip())


# ─────────────────────────────────────────────────────────────────────
# 解析结果
# ─────────────────────────────────────────────────────────────────────
@dataclass
class ParsedWorkbook:
    """回读产物。`errors` 非空即说明版式层面已经不合契约。"""

    plan: SchedulePlan | None = None
    projection: ThreeTableProjection = field(default_factory=ThreeTableProjection)
    sheet_names: tuple[str, ...] = ()
    errors: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class _Block7Row:
    sortie_id: str
    date: date_type
    takeoff: time_type
    aircraft_id: str
    aircraft_type: str
    runway_id: str
    airspace_id: str
    is_recurrent: bool


def _name_resolver(
    ctx: ValidationContext | None, plan: SchedulePlan | None
) -> tuple[dict[str, str], list[str]]:
    """姓名 → `person_id`。优先用人员表；重名不猜，直接报错。"""
    errors: list[str] = []
    mapping: dict[str, str] = {}
    seen: dict[str, set[str]] = {}
    if ctx is not None:
        for p in ctx.sorted_persons():
            seen.setdefault(p.name, set()).add(p.person_id)
    elif plan is not None:
        for s in plan.sorties:
            for c in s.crew:
                seen.setdefault(c.name, set()).add(c.person_id)
    for name, ids in sorted(seen.items()):
        if len(ids) > 1:
            errors.append(f"姓名 {name} 对应多个编号 {sorted(ids)}，Sheet 1~3 只写姓名时无法反解")
            continue
        mapping[name] = next(iter(ids))
    return mapping, errors


# ─────────────────────────────────────────────────────────────────────
# Sheet 4
# ─────────────────────────────────────────────────────────────────────
def _split_blocks(rows: Sequence[Sequence[Any]]) -> dict[str, list[list[Any]]]:
    blocks: dict[str, list[list[Any]]] = {}
    current: str | None = None
    for row in rows:
        head = _text(row[0]) if row else ""
        if head in BLOCK_TITLES:
            current = head
            blocks[current] = []
            continue
        if current is not None:
            blocks[current].append(list(row))
    return blocks


def _meta_map(block: Sequence[Sequence[Any]]) -> dict[str, str]:
    return {_text(r[0]): _text(r[1]) for r in block if len(r) > 1 and _text(r[0])}


def _parse_person_cell(text: str, where: str) -> str:
    """Sheet 4 的人员列写作 `何超(P08)` —— Sheet 4 无版式基准，故直接带编号。"""
    m = re.match(r"^.*\((P\d+)\)$", text.strip())
    if not m:
        raise WorkbookFormatError(f"{where} 的人员单元格应为 `姓名(编号)` 形态，实际 {text!r}")
    return m.group(1)


def _parse_block7(
    block: Sequence[Sequence[Any]],
) -> dict[tuple[date_type, time_type, str], _Block7Row]:
    if not block or tuple(_text(c) for c in block[0][: len(BLOCK7_HEADERS)]) != BLOCK7_HEADERS:
        raise WorkbookFormatError(
            f"区块7 表头必须逐字为 {BLOCK7_HEADERS}，实际 "
            f"{tuple(_text(c) for c in (block[0] if block else ()))}"
        )
    out: dict[tuple[date_type, time_type, str], _Block7Row] = {}
    for row in block[1:]:
        if _is_blank(row):
            continue
        day = _parse_date(row[1], "区块7 日期列")
        takeoff = _parse_time(_require_time_text(row[2], "区块7 起飞列"))
        mark = _text(row[7])
        parsed = _Block7Row(
            sortie_id=_text(row[0]),
            date=day,
            takeoff=takeoff,
            aircraft_id=_text(row[3]),
            aircraft_type=_text(row[4]),
            runway_id=_text(row[5]),
            airspace_id=_text(row[6]),
            is_recurrent=mark == RECURRENT_MARK,
        )
        if mark not in (RECURRENT_MARK, NOT_RECURRENT_MARK):
            raise WorkbookFormatError(
                f"区块7 复训标记列取值只能是 {RECURRENT_MARK!r} 或 {NOT_RECURRENT_MARK!r}，实际 {mark!r}"
            )
        out[(day, takeoff, parsed.aircraft_id)] = parsed
    return out


def _parse_debts(block: Sequence[Sequence[Any]]) -> list[TrainingDebt]:
    if not block:
        return []
    if tuple(_text(c) for c in block[0][: len(BLOCK3_HEADERS)]) != BLOCK3_HEADERS:
        raise WorkbookFormatError(f"区块3 表头必须逐字为 {BLOCK3_HEADERS}")
    debts: list[TrainingDebt] = []
    for row in block[1:]:
        if _is_blank(row) or _text(row[9]) in ("", "—"):
            continue  # 只有带松弛档的行才是 TrainingDebt，其余是进度展示行
        debts.append(
            TrainingDebt(
                person_id=_parse_person_cell(_text(row[0]), "区块3 人员列"),
                mission_id=_text(row[1]),
                required=int(_text(row[4])),
                scheduled=int(_text(row[5])),
                debt=int(_text(row[6])),
                relaxed_by=_text(row[9]),  # type: ignore[arg-type]
            )
        )
    return debts


def _parse_blocked(block: Sequence[Sequence[Any]]) -> list[BlockedItem]:
    if not block:
        return []
    if tuple(_text(c) for c in block[0][: len(BLOCK4_HEADERS)]) != BLOCK4_HEADERS:
        raise WorkbookFormatError(f"区块4 表头必须逐字为 {BLOCK4_HEADERS}")
    items: list[BlockedItem] = []
    for row in block[1:]:
        if _is_blank(row):
            continue
        missing = [m for m in _text(row[3]).replace(",", "、").split("、") if m and m != "—"]
        items.append(
            BlockedItem(
                person_id=_parse_person_cell(_text(row[0]), "区块4 人员列"),
                mission_id=_text(row[1]),
                reason=_text(row[2]),
                missing_prereqs=missing,
            )
        )
    return items


# ─────────────────────────────────────────────────────────────────────
# Sheet 1~3
# ─────────────────────────────────────────────────────────────────────
@dataclass(frozen=True)
class _Sheet1Row:
    weekday: str
    takeoff: time_type
    landing: time_type
    aircraft_id: str
    mission_id: str
    mission_name: str
    crew: tuple[tuple[str, str], ...]  # (姓名, 角色)


def _parse_sheet1(rows: Sequence[Sequence[Any]]) -> list[_Sheet1Row]:
    out: list[_Sheet1Row] = []
    weekday: str | None = None
    expect_header = False
    seen_days: list[str] = []
    for n, row in enumerate(rows, start=1):
        if _is_blank(row):
            continue
        head = _text(row[0])
        if head in WEEKDAY_ORDER and all(_text(c) == "" for c in row[1:]):
            weekday = head
            seen_days.append(head)
            expect_header = True
            continue
        if expect_header:
            if tuple(_text(c) for c in row[: len(SHEET1_HEADERS)]) != SHEET1_HEADERS:
                raise WorkbookFormatError(
                    f"Sheet 1 第 {n} 行表头必须逐字为 {SHEET1_HEADERS}，实际 "
                    f"{tuple(_text(c) for c in row[: len(SHEET1_HEADERS)])}"
                )
            expect_header = False
            continue
        if weekday is None:
            raise WorkbookFormatError(f"Sheet 1 第 {n} 行出现在任何星期分组之前")
        name, mid, _ = parse_mission_cell(_text(row[3]), f"Sheet 1 第 {n} 行")
        out.append(
            _Sheet1Row(
                weekday=weekday,
                takeoff=_parse_time(_require_time_text(row[0], f"Sheet 1 第 {n} 行起飞列")),
                landing=_parse_time(_require_time_text(row[1], f"Sheet 1 第 {n} 行着陆列")),
                aircraft_id=_text(row[2]),
                mission_id=mid,
                mission_name=name,
                crew=parse_sheet1_crew(_text(row[4]), f"Sheet 1 第 {n} 行"),
            )
        )
    order = [d for d in WEEKDAY_ORDER if d in seen_days]
    if seen_days != order:
        raise WorkbookFormatError(f"Sheet 1 的星期分组顺序应为周一~周日，实际 {seen_days}")
    if len(set(seen_days)) != len(seen_days):
        raise WorkbookFormatError(f"Sheet 1 的星期分组重复出现：{seen_days}")
    return out


@dataclass(frozen=True)
class _GroupedRow:
    group: str
    weekday: str
    takeoff: time_type
    landing: time_type | None
    mission_id: str
    mission_name: str
    payload: tuple[str, ...]


def _parse_grouped_sheet(
    rows: Sequence[Sequence[Any]], headers: tuple[str, ...], sheet: str
) -> list[_GroupedRow]:
    """Sheet 2/3 的「一级分组 → 二级分组（星期）→ 数据行」结构。"""
    out: list[_GroupedRow] = []
    group: str | None = None
    weekday: str | None = None
    weekdays_in_group: list[str] = []
    groups: list[str] = []
    expect_header = False
    for n, row in enumerate(rows, start=1):
        if _is_blank(row):
            continue
        col_a, col_b = _text(row[0]), _text(row[1])
        if col_a and not col_b and all(_text(c) == "" for c in row[2:]):
            group = col_a
            groups.append(group)
            weekdays_in_group = []
            weekday = None
            expect_header = True
            continue
        if expect_header:
            if tuple(_text(c) for c in row[: len(headers)]) != headers:
                raise WorkbookFormatError(
                    f"{sheet} 第 {n} 行表头必须逐字为 {headers}，实际 "
                    f"{tuple(_text(c) for c in row[: len(headers)])}"
                )
            expect_header = False
            continue
        if col_b and not col_a and all(_text(c) == "" for c in row[2:]):
            if col_b not in WEEKDAY_ORDER:
                raise WorkbookFormatError(f"{sheet} 第 {n} 行的二级分组 {col_b!r} 不是合法星期")
            weekday = col_b
            weekdays_in_group.append(col_b)
            order = [d for d in WEEKDAY_ORDER if d in weekdays_in_group]
            if weekdays_in_group != order:
                raise WorkbookFormatError(
                    f"{sheet} 分组 {group!r} 的星期顺序应为周一~周日，实际 {weekdays_in_group}"
                )
            continue
        if group is None or weekday is None:
            raise WorkbookFormatError(f"{sheet} 第 {n} 行出现在分组标题之前")
        time_cell = _text(row[2])
        if "-" in time_cell:
            t0_text, t1_text = (part.strip() for part in time_cell.split("-", 1))
            takeoff = _parse_time(_require_time_text(t0_text, f"{sheet} 第 {n} 行时间列"))
            landing: time_type | None = _parse_time(
                _require_time_text(t1_text, f"{sheet} 第 {n} 行时间列")
            )
        else:
            takeoff = _parse_time(_require_time_text(row[2], f"{sheet} 第 {n} 行时间列"))
            landing = None
        mission_name, mid, _asp = parse_mission_cell(_text(row[3]), f"{sheet} 第 {n} 行")
        out.append(
            _GroupedRow(
                group=group,
                weekday=weekday,
                takeoff=takeoff,
                landing=landing,
                mission_id=mid,
                mission_name=mission_name,
                payload=parse_paren_pair(_text(row[4]), f"{sheet} 第 {n} 行"),
            )
        )
    if len(set(groups)) != len(groups):
        raise WorkbookFormatError(f"{sheet} 的一级分组重复出现：{groups}")
    return out


# ─────────────────────────────────────────────────────────────────────
# 回读主流程
# ─────────────────────────────────────────────────────────────────────
def parse_workbook(path: Path, *, name_map: Mapping[str, str]) -> ParsedWorkbook:
    """把 xlsx 反解为 `SchedulePlan` + 三表投影。异常收敛为 `errors`。

    `name_map` 是「姓名 → person_id」—— Sheet 1~3 按版式基准只写姓名，反解必须
    靠外部的人员表（见 `_name_resolver`）。
    """
    result = ParsedWorkbook()
    wb = load_workbook(path, data_only=True)
    result.sheet_names = tuple(wb.sheetnames)
    if result.sheet_names != SHEET_ORDER:
        result.errors.append(f"工作表名与顺序必须为 {SHEET_ORDER}，实际 {result.sheet_names}")
        return result

    try:
        s4_rows = _row_values(wb[SHEET_ORDER[3]], 10)
        blocks = _split_blocks(s4_rows)
        missing_blocks = [t for t in BLOCK_TITLES if t not in blocks]
        if missing_blocks:
            raise WorkbookFormatError(f"Sheet 4 缺少区块：{missing_blocks}")
        meta = _meta_map(blocks[BLOCK_TITLES[0]])
        missing_meta = [label for label in REQUIRED_META_LABELS if label not in meta]
        if missing_meta:
            raise WorkbookFormatError(f"区块1 缺少反解所需字段：{missing_meta}")
        cover = _RANGE_RE.match(meta[META_COVER])
        if not cover:
            raise WorkbookFormatError(f"区块1「{META_COVER}」应为 `YYYY-MM-DD ~ YYYY-MM-DD`")
        week_start = date_type.fromisoformat(cover.group(1))
        week_end = date_type.fromisoformat(cover.group(2))
        tier_m = _TIER_RE.search(meta[META_TIER])
        if not tier_m:
            raise WorkbookFormatError(f"区块1「{META_TIER}」里找不到档位数字：{meta[META_TIER]!r}")
        switches = {
            k.strip(): v.strip()
            for k, _, v in (
                item.partition(SWITCH_KV)
                for item in meta[META_SWITCHES].split(SWITCH_SEP)
                if item.strip()
            )
        }
        block7 = _parse_block7(blocks[BLOCK_TITLES[6]])
        debts = _parse_debts(blocks[BLOCK_TITLES[2]])
        blocked = _parse_blocked(blocks[BLOCK_TITLES[3]])

        s1 = _parse_sheet1(_row_values(wb[SHEET_ORDER[0]], len(SHEET1_HEADERS)))
        s2 = _parse_grouped_sheet(
            _row_values(wb[SHEET_ORDER[1]], len(SHEET2_HEADERS)), SHEET2_HEADERS, "Sheet 2"
        )
        s3 = _parse_grouped_sheet(
            _row_values(wb[SHEET_ORDER[2]], len(SHEET3_HEADERS)), SHEET3_HEADERS, "Sheet 3"
        )
    except WorkbookFormatError as exc:
        result.errors.append(str(exc))
        return result

    sorties: list[Sortie] = []
    for row in s1:
        day = week_start + _weekday_offset(row.weekday)
        key = (day, row.takeoff, row.aircraft_id)
        b7 = block7.get(key)
        if b7 is None:
            result.errors.append(f"区块7 缺少 {day} {row.takeoff} {row.aircraft_id} 对应的明细行")
            continue
        crew: list[CrewMember] = []
        for name, role in row.crew:
            person_id = name_map.get(name)
            if person_id is None:
                result.errors.append(f"人员表里找不到姓名 {name!r}（{b7.sortie_id}）")
                continue
            crew.append(CrewMember(person_id=person_id, name=name, role=role))  # type: ignore[arg-type]
        try:
            sorties.append(
                Sortie(
                    sortie_id=b7.sortie_id,
                    date=day,
                    weekday=row.weekday,  # type: ignore[arg-type]
                    takeoff=row.takeoff,
                    landing=row.landing,
                    mission_id=row.mission_id,
                    mission_name=row.mission_name,
                    airspace_id=b7.airspace_id,
                    aircraft_id=row.aircraft_id,
                    runway_id=b7.runway_id,
                    is_recurrent=b7.is_recurrent,
                    crew=crew,
                )
            )
        except ValueError as exc:
            result.errors.append(f"{b7.sortie_id} 反解后不满足 Sortie 契约：{exc}")

    try:
        plan = SchedulePlan(
            plan_id=meta[META_PLAN_ID],
            iso_week=meta[META_ISO_WEEK],
            week_start=week_start,
            week_end=week_end,
            snapshot_id=meta[META_SNAPSHOT],
            ruleset_version=meta[META_RULESET],
            semantics_version=meta[META_SEMANTICS],
            semantics_switches=switches,
            runway_model=meta[META_RUNWAY_MODEL].split("（")[0].strip(),  # type: ignore[arg-type]
            relaxation_tier=int(tier_m.group(1)),
            sorties=sorties,
            debts=debts,
            blocked_items=blocked,
            content_sha256=meta[META_SHA],
        )
    except ValueError as exc:
        result.errors.append(f"区块1 反解后不满足 SchedulePlan 契约：{exc}")
        return result

    result.plan = plan
    result.projection = _projection_from_sheets(plan, s2, s3, name_map, result.errors)
    return result


def _weekday_offset(weekday: str) -> timedelta:
    return timedelta(days=WEEKDAY_ORDER.index(weekday))


def _projection_from_sheets(
    plan: SchedulePlan,
    s2: Sequence[_GroupedRow],
    s3: Sequence[_GroupedRow],
    name_map: Mapping[str, str],
    errors: list[str],
) -> ThreeTableProjection:
    """用 Sheet 1（已并入 plan）/ Sheet 2 / Sheet 3 三张表各自的内容建投影。

    **不是从 plan 投影三次** —— 那样三表交叉一致性就成了自证。Sheet 2/3 的行按
    (日期, 起飞, 机号) 与 Sheet 1 的架次对齐，再逐字段比对。
    """
    by_key = {(s.date, s.takeoff, s.aircraft_id): s for s in plan.sorties}
    by_day: dict[date_type, list[SortieRow]] = {}
    for s in plan.sorties:
        by_day.setdefault(s.date, []).append(
            SortieRow(
                sortie_id=s.sortie_id,
                date=s.date,
                weekday=s.weekday,
                takeoff=s.takeoff,
                landing=s.landing,
                mission_id=s.mission_id,
                mission_name=s.mission_name,
                aircraft_id=s.aircraft_id,
                crew=tuple(sorted((c.person_id, c.role) for c in s.crew)),
            )
        )

    by_person: dict[str, list[SortieRow]] = {}
    for row in s2:
        person_id = name_map.get(row.group)
        if person_id is None:
            errors.append(f"Sheet 2 的分组 {row.group!r} 在人员表里找不到")
            continue
        day = plan.week_start + _weekday_offset(row.weekday)
        aircraft_id, role = (*row.payload, "", "")[:2]
        src = by_key.get((day, row.takeoff, aircraft_id))
        if src is None:
            errors.append(
                f"Sheet 2 的 {row.group} {row.weekday} {row.takeoff} 在分日表里找不到对应架次"
            )
            continue
        by_person.setdefault(person_id, []).append(
            SortieRow(
                sortie_id=src.sortie_id,
                date=day,
                weekday=row.weekday,
                takeoff=row.takeoff,
                landing=row.landing or src.landing,
                mission_id=row.mission_id,
                mission_name=row.mission_name,
                aircraft_id=aircraft_id,
                crew=tuple(sorted((c.person_id, c.role) for c in src.crew)),
            )
        )
        if (person_id, role) not in {(c.person_id, c.role) for c in src.crew}:
            errors.append(
                f"Sheet 2：{row.group}({person_id}) 在 {src.sortie_id} 上的角色写作 {role!r}，"
                f"与分日表的 {[(c.person_id, c.role) for c in src.crew]} 不一致"
            )

    by_aircraft: dict[str, list[SortieRow]] = {}
    for row in s3:
        day = plan.week_start + _weekday_offset(row.weekday)
        src = by_key.get((day, row.takeoff, row.group))
        if src is None:
            errors.append(
                f"Sheet 3 的 {row.group} {row.weekday} {row.takeoff} 在分日表里找不到对应架次"
            )
            continue
        # Sheet 3 的机组列只有姓名：双人按 (教员, 学员) 顺序，单人按复训标记定角色
        names = list(row.payload)
        solo_role = "复训" if src.is_recurrent else "单飞"
        roles = ["教员", "学员"] if len(names) == 2 else [solo_role]
        crew: list[tuple[str, str]] = []
        for name, role in zip(names, roles, strict=False):
            person_id = name_map.get(name)
            if person_id is None:
                errors.append(f"Sheet 3 的机组姓名 {name!r} 在人员表里找不到")
                continue
            crew.append((person_id, role))
        by_aircraft.setdefault(row.group, []).append(
            SortieRow(
                sortie_id=src.sortie_id,
                date=day,
                weekday=row.weekday,
                takeoff=row.takeoff,
                landing=src.landing,
                mission_id=row.mission_id,
                mission_name=row.mission_name,
                aircraft_id=row.group,
                crew=tuple(sorted(crew)),
            )
        )

    def _freeze(group: dict[Any, list[SortieRow]]) -> dict[Any, tuple[SortieRow, ...]]:
        return {
            k: tuple(sorted(v, key=lambda r: (r.date, r.takeoff, r.sortie_id)))
            for k, v in sorted(group.items(), key=lambda kv: str(kv[0]))
        }

    return ThreeTableProjection(
        by_day=_freeze(by_day), by_person=_freeze(by_person), by_aircraft=_freeze(by_aircraft)
    )


# ─────────────────────────────────────────────────────────────────────
# 深度比对
# ─────────────────────────────────────────────────────────────────────
def _normalized(plan: SchedulePlan) -> dict[str, Any]:
    """深度比对前把三个列表排成确定序 —— 顺序不是契约的一部分，内容才是。"""
    dumped: dict[str, Any] = plan.model_dump(mode="json")
    dumped["sorties"] = sorted(dumped["sorties"], key=lambda s: s["sortie_id"])
    dumped["debts"] = sorted(dumped["debts"], key=lambda d: (d["person_id"], d["mission_id"]))
    dumped["blocked_items"] = sorted(
        dumped["blocked_items"], key=lambda b: (b["person_id"], b["mission_id"])
    )
    return dumped


def deep_diff(source: Any, parsed: Any, path: str = "") -> list[str]:
    """逐字段深度比对，返回差异路径。**必须为空**（v6 §4.3）。"""
    where = path or "<root>"
    if isinstance(source, dict) and isinstance(parsed, dict):
        diffs: list[str] = []
        for key in sorted(set(source) | set(parsed)):
            if key not in source:
                diffs.append(f"{where}.{key}: 源对象没有该字段，回读得到 {parsed[key]!r}")
            elif key not in parsed:
                diffs.append(f"{where}.{key}: 回读丢失（源为 {source[key]!r}）")
            else:
                diffs.extend(deep_diff(source[key], parsed[key], f"{path}.{key}" if path else key))
        return diffs
    if isinstance(source, list) and isinstance(parsed, list):
        if len(source) != len(parsed):
            return [f"{where}: 长度 {len(source)} → 回读 {len(parsed)}"]
        out: list[str] = []
        for i, (a, b) in enumerate(zip(source, parsed, strict=True)):
            out.extend(deep_diff(a, b, f"{path}[{i}]"))
        return out
    if source != parsed:
        return [f"{where}: {source!r} → 回读 {parsed!r}"]
    return []


def verify_workbook(
    path: Path, plan: SchedulePlan, *, ctx: ValidationContext | None = None
) -> SchemaCheckReport:
    """闸门3：回读 xlsx，断言版式、单元格类型、分组结构与**深度相等**。

    `ctx` 用于把 Sheet 1~3 里的姓名映射回 `person_id`（版式基准里人员只有姓名），
    并在 S-11 开关为 on 时检查区块6 的授权改写声明是否在场。
    """
    name_map, name_errors = _name_resolver(ctx, plan)
    parsed = parse_workbook(path, name_map=name_map)
    diff = [*name_errors, *parsed.errors]
    if parsed.plan is not None:
        diff.extend(deep_diff(_normalized(plan), _normalized(parsed.plan)))
        diff.extend(check_cross_table_consistency(parsed.projection))
    if ctx is not None and ctx.semantics.s11_enabled:
        diff.extend(_check_s11_declaration(path))
    return SchemaCheckReport(
        passed=not diff,
        diff=diff,
        workbook_path=str(path),
        sheet_names=list(parsed.sheet_names),
    )


def _check_s11_declaration(path: Path) -> list[str]:
    """v6 §10.4 区块6 的强制项：只要 S-11 开关为 on，授权改写声明必须在场。"""
    wb = load_workbook(path, data_only=True)
    if SHEET_ORDER[3] not in wb.sheetnames:
        return []
    blocks = _split_blocks(_row_values(wb[SHEET_ORDER[3]], 10))
    rows = blocks.get(BLOCK_TITLES[5], [])
    if not any(_text(r[0]) == "授权改写声明" and _text(r[1]) for r in rows if len(r) > 1):
        return ["区块6 缺少「授权改写声明」行（S-11 开关为 on 时是强制项，v6 §10.4）"]
    return []


__all__ = [
    "BLOCK2_HEADERS",
    "BLOCK3_HEADERS",
    "BLOCK4_HEADERS",
    "BLOCK7_HEADERS",
    "BLOCK_TITLES",
    "META_COVER",
    "META_ISO_WEEK",
    "META_PLAN_ID",
    "META_RULESET",
    "META_RUNWAY_MODEL",
    "META_SEMANTICS",
    "META_SHA",
    "META_SNAPSHOT",
    "META_SWITCHES",
    "META_TIER",
    "NOT_RECURRENT_MARK",
    "RECURRENT_MARK",
    "REQUIRED_META_LABELS",
    "ROLE_SUFFIX",
    "SHEET1_HEADERS",
    "SHEET2_HEADERS",
    "SHEET3_HEADERS",
    "SHEET_ORDER",
    "SUFFIX_ROLE",
    "SWITCH_KV",
    "SWITCH_SEP",
    "ParsedWorkbook",
    "WorkbookFormatError",
    "deep_diff",
    "parse_mission_cell",
    "parse_paren_pair",
    "parse_sheet1_crew",
    "parse_workbook",
    "verify_workbook",
]
