"""闸门3 的单元测试：Excel 回读与深度相等断言（v6 §4.3）。

xlsx 由 `tests/fixtures/workbook_builder.py` 按 `backend.validator.workbook` 的
回读契约手工写出（M3 的正式写出模块还没有）。**关键不是「能通过」，而是「篡改
之后必须通不过」** —— 一个只会返回 `passed=True` 的回读器同样"全绿"，那种绿是
假的，所以下面每个用例都在产物或源对象上做单点篡改。
"""

from __future__ import annotations

from datetime import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.schemas.plan import SchedulePlan
from backend.validator.context import ValidationContext
from backend.validator.workbook import (
    BLOCK_TITLES,
    META_SWITCHES,
    SHEET_ORDER,
    WorkbookFormatError,
    deep_diff,
    parse_mission_cell,
    parse_paren_pair,
    parse_sheet1_crew,
    parse_workbook,
    verify_workbook,
)
from tests.fixtures.validator_facts import (
    AIRCRAFT_TYPE,
    baseline_context,
    compliant_plan,
    debt,
    make_plan,
    make_sortie,
)
from tests.fixtures.workbook_builder import write_workbook


@pytest.fixture(scope="module")
def ctx() -> ValidationContext:
    return baseline_context()


@pytest.fixture
def plan() -> SchedulePlan:
    return compliant_plan()


@pytest.fixture
def book(tmp_path: Path, plan: SchedulePlan) -> Path:
    return write_workbook(plan, tmp_path / "plan.xlsx", aircraft_types=AIRCRAFT_TYPE)


# ─────────────────────────────────────────────────────────────────────
# 正向：写出 → 回读 → 深度相等
# ─────────────────────────────────────────────────────────────────────
def test_roundtrip_is_deeply_equal(book: Path, plan: SchedulePlan, ctx: ValidationContext) -> None:
    report = verify_workbook(book, plan, ctx=ctx)
    assert report.passed, report.diff
    assert report.sheet_names == list(SHEET_ORDER)
    assert report.workbook_path == str(book)


def test_roundtrip_recovers_every_sortie_field(book: Path, ctx: ValidationContext) -> None:
    """`runway_id` 与 `is_recurrent` 只能从区块7 反解（v6 §10.4 / §4.3）。"""
    parsed = parse_workbook(book, name_map={p.name: p.person_id for p in ctx.sorted_persons()})
    assert parsed.errors == []
    assert parsed.plan is not None
    runways = {s.sortie_id: s.runway_id for s in parsed.plan.sorties}
    assert runways["S000002"] == "RWY-2"
    assert runways["S000001"] == "RWY-1"
    assert all(s.is_recurrent is False for s in parsed.plan.sorties)


def test_roundtrip_handles_recurrent_sortie(tmp_path: Path, ctx: ValidationContext) -> None:
    """S-11 的复训架次：角色「复训」+ 区块7 的复训标记都要能往返。"""
    recurrent = make_sortie(
        "S000910", 3, "07:20", "missionC-1", "AC61", (("P04", "复训"),), is_recurrent=True
    )
    one = make_plan([recurrent])
    path = write_workbook(one, tmp_path / "recurrent.xlsx", aircraft_types=AIRCRAFT_TYPE)
    report = verify_workbook(path, one, ctx=ctx)
    assert report.passed, report.diff


def test_roundtrip_carries_debts_and_blocked_items(tmp_path: Path, ctx: ValidationContext) -> None:
    relaxed = compliant_plan(
        relaxation_tier=1,
        debts=[debt("P05", "missionC-2", required=1, scheduled=0, relaxed_by="TIER1")],
    )
    path = write_workbook(relaxed, tmp_path / "relaxed.xlsx", aircraft_types=AIRCRAFT_TYPE)
    report = verify_workbook(path, relaxed, ctx=ctx)
    assert report.passed, report.diff
    parsed = parse_workbook(path, name_map={p.name: p.person_id for p in ctx.sorted_persons()})
    assert parsed.plan is not None
    assert len(parsed.plan.debts) == 1
    assert len(parsed.plan.blocked_items) == 7  # v6 §1.4.2 的 7 条


def test_verify_workbook_works_without_context(book: Path, plan: SchedulePlan) -> None:
    """没有人员表时用方案自带的机组姓名兜底 —— 但这只够做格式比对。"""
    assert verify_workbook(book, plan).passed


# ─────────────────────────────────────────────────────────────────────
# 反向：篡改之后必须通不过
# ─────────────────────────────────────────────────────────────────────
def test_detects_wrong_sheet_names(
    tmp_path: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    path = write_workbook(
        plan,
        tmp_path / "bad_sheets.xlsx",
        aircraft_types=AIRCRAFT_TYPE,
        sheet_names=("Sheet1", "Sheet2", "Sheet3", "Sheet4"),
    )
    report = verify_workbook(path, plan, ctx=ctx)
    assert not report.passed
    assert any("工作表名与顺序必须为" in d for d in report.diff)


def test_detects_excel_serial_time_cells(
    tmp_path: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    """★ v6 §4.3 点名的断言：时间列必须是 `HH:MM` 文本，不是 Excel 序列号。"""
    path = write_workbook(
        plan, tmp_path / "native_time.xlsx", aircraft_types=AIRCRAFT_TYPE, time_as_native=True
    )
    report = verify_workbook(path, plan, ctx=ctx)
    assert not report.passed
    assert any("必须写成 HH:MM 文本" in d for d in report.diff)


def test_detects_content_drift_between_plan_and_workbook(
    book: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    """源对象改一个字段、产物没跟着改 → 深度比对必须指出是哪个字段。"""
    drifted = plan.model_copy(
        update={
            "sorties": [
                plan.sorties[0].model_copy(update={"runway_id": "RWY-2"}),
                *plan.sorties[1:],
            ]
        }
    )
    report = verify_workbook(book, drifted, ctx=ctx)
    assert not report.passed
    assert any("runway_id" in d for d in report.diff)


def test_detects_missing_block7_row(book: Path, plan: SchedulePlan, ctx: ValidationContext) -> None:
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[3]]
    target = next(
        row for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=1).value == "S000001"
    )
    ws.delete_rows(target)
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("区块7 缺少" in d for d in report.diff)


def test_detects_missing_semantics_switches_row(
    book: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    """`semantics_switches` 参与 `content_sha256`，不落表就反解不回来。"""
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[3]]
    target = next(
        row for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=1).value == META_SWITCHES
    )
    ws.delete_rows(target)
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("区块1 缺少反解所需字段" in d for d in report.diff)


def test_detects_tampered_header(book: Path, plan: SchedulePlan, ctx: ValidationContext) -> None:
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[0]]
    ws.cell(row=2, column=4).value = "课目"  # 应为「课目（空域）」
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("表头必须逐字为" in d for d in report.diff)


def test_detects_wrong_weekday_group_order(
    book: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[0]]
    ws.cell(row=1, column=1).value = "周日"  # 第一组本应是周一
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("星期分组顺序" in d for d in report.diff)


def test_detects_role_drift_in_person_sheet(
    book: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    """Sheet 2 的「飞机/角色」写错 → 三表交叉一致性必须拦下。"""
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[1]]
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=5)
        if isinstance(cell.value, str) and cell.value.endswith("/学员)"):
            cell.value = cell.value.replace("/学员)", "/单飞)")
            break
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("角色写作" in d for d in report.diff)


def test_detects_row_drift_in_aircraft_sheet(
    book: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[2]]
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if cell.value == "06:00":
            cell.value = "06:05"
            break
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("在分日表里找不到对应架次" in d for d in report.diff)


def test_detects_missing_s11_declaration(
    book: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    """v6 §10.4 区块6：S-11 开关为 on 时授权改写声明是强制项。"""
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[3]]
    target = next(
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=1).value == "授权改写声明"
    )
    ws.delete_rows(target)
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("授权改写声明" in d for d in report.diff)
    # 没有 ctx 就读不到 S-11 开关，这一项跳过 —— 其余比对照常通过
    assert verify_workbook(book, plan).passed


def test_detects_missing_block(book: Path, plan: SchedulePlan, ctx: ValidationContext) -> None:
    wb = load_workbook(book)
    ws = wb[SHEET_ORDER[3]]
    target = next(
        row
        for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=1).value == BLOCK_TITLES[4]
    )
    ws.cell(row=target, column=1).value = "区块5 资源利用"  # 少了「 · 」
    wb.save(book)
    report = verify_workbook(book, plan, ctx=ctx)
    assert not report.passed
    assert any("Sheet 4 缺少区块" in d for d in report.diff)


def test_duplicate_person_names_are_reported_not_guessed(
    book: Path, plan: SchedulePlan, ctx: ValidationContext
) -> None:
    """重名时姓名反解不出编号 —— 如实报错，不猜。"""
    import dataclasses

    persons = dict(ctx.persons)
    persons["P07"] = dataclasses.replace(persons["P07"], name="罗磊")
    twins = dataclasses.replace(ctx, persons=persons)
    report = verify_workbook(book, plan, ctx=twins)
    assert not report.passed
    assert any("对应多个编号" in d for d in report.diff)


# ─────────────────────────────────────────────────────────────────────
# 拼接格式与深度比对的单元测试
# ─────────────────────────────────────────────────────────────────────
def test_parse_mission_cell() -> None:
    assert parse_mission_cell("本场起落航线 (missionA-1)（Small Area A）", "x") == (
        "本场起落航线",
        "missionA-1",
        "Small Area A",
    )
    assert parse_mission_cell("仪表飞行 (missionC-1)", "x") == ("仪表飞行", "missionC-1", None)
    with pytest.raises(WorkbookFormatError, match="不合拼接格式"):
        parse_mission_cell("仪表飞行", "x")


def test_parse_sheet1_crew() -> None:
    assert parse_sheet1_crew("孙军教，陈伟学", "x") == (("孙军", "教员"), ("陈伟", "学员"))
    assert parse_sheet1_crew("何超单", "x") == (("何超", "单飞"),)
    assert parse_sheet1_crew("刘斌训", "x") == (("刘斌", "复训"),)
    with pytest.raises(WorkbookFormatError, match="缺少角色后缀"):
        parse_sheet1_crew("孙军", "x")


def test_parse_paren_pair() -> None:
    assert parse_paren_pair("(AC49/学员)", "x") == ("AC49", "学员")
    assert parse_paren_pair("(高超/罗磊)", "x") == ("高超", "罗磊")
    with pytest.raises(WorkbookFormatError, match="应为"):
        parse_paren_pair("AC49/学员", "x")


def test_deep_diff_reports_paths() -> None:
    assert deep_diff({"a": 1}, {"a": 1}) == []
    assert deep_diff({"a": 1}, {"a": 2}) == ["a: 1 → 回读 2"]
    assert deep_diff({"a": [1, 2]}, {"a": [1]}) == ["a: 长度 2 → 回读 1"]
    assert deep_diff({"a": 1}, {}) == ["<root>.a: 回读丢失（源为 1）"]
    assert deep_diff({}, {"b": 3}) == ["<root>.b: 源对象没有该字段，回读得到 3"]
    nested = deep_diff({"s": [{"t": time(6, 0).isoformat()}]}, {"s": [{"t": "06:30:00"}]})
    assert nested == ["s[0].t: '06:00:00' → 回读 '06:30:00'"]
