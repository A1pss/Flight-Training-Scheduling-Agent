"""四表渲染的单元测试（v6 §10.1~§10.4 + §4.3）。

不连库、不跑求解器 —— 实体与方案取 `tests/fixtures/validator_facts.py` 的手工样本。
"""

from __future__ import annotations

from datetime import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.report.excel import (
    DASH,
    S11_DECLARATION_LABEL,
    fmt_crew_sheet1,
    fmt_crew_sheet3,
    fmt_mission,
    fmt_switches,
    fmt_time,
    minimum_required,
    render_workbook,
)
from backend.report.template import build_template, read_template_headers, template_contract
from backend.schemas.plan import CrewMember
from backend.validator.workbook import (
    BLOCK_TITLES,
    REQUIRED_META_LABELS,
    SHEET1_HEADERS,
    SHEET2_HEADERS,
    SHEET3_HEADERS,
    SHEET_ORDER,
    deep_diff,
    parse_workbook,
    verify_workbook,
)
from tests.fixtures.report_bundle import RECURRENT_SORTIE_ID, sample_bundle
from tests.fixtures.validator_facts import PERSON_NAME


@pytest.fixture(scope="module")
def rendered(tmp_path_factory: pytest.TempPathFactory) -> tuple[Path, object]:
    bundle = sample_bundle()
    path = tmp_path_factory.mktemp("report") / "plan.xlsx"
    render_workbook(path, bundle, readback_passed=True)
    return path, bundle


# ─────────────────────────────────────────────────────────────────────
# 拼接格式（与回读解析器互为逆运算）
# ─────────────────────────────────────────────────────────────────────
def test_time_format_is_zero_padded_hh_mm() -> None:
    assert fmt_time(time(6, 0)) == "06:00"
    assert fmt_time(time(16, 20)) == "16:20"


def test_mission_cell_matches_layout_baseline() -> None:
    assert (
        fmt_mission("本场起落航线", "missionA-1", "Small Area A")
        == "本场起落航线 (missionA-1)（Small Area A）"
    )
    assert fmt_mission("仪表飞行", "missionC-1") == "仪表飞行 (missionC-1)"


def test_crew_cell_uses_four_role_suffixes() -> None:
    """教「教」/ 学「学」/ 单飞「单」/ 复训「训」 —— 「训」是 v6 新增（§10.1）。"""
    dual = [
        CrewMember(person_id="P01", name="孙军", role="教员"),
        CrewMember(person_id="P07", name="陈伟", role="学员"),
    ]
    assert fmt_crew_sheet1(dual) == "孙军教，陈伟学"
    assert fmt_crew_sheet1([CrewMember(person_id="P08", name="何超", role="单飞")]) == "何超单"
    assert fmt_crew_sheet1([CrewMember(person_id="P04", name="刘斌", role="复训")]) == "刘斌训"
    assert fmt_crew_sheet3(dual) == "(孙军/陈伟)"


def test_crew_order_is_instructor_first() -> None:
    """机组顺序固定「教员，学员」——回读侧按位置定角色（Sheet 3 只有姓名）。"""
    reversed_crew = [
        CrewMember(person_id="P07", name="陈伟", role="学员"),
        CrewMember(person_id="P01", name="孙军", role="教员"),
    ]
    assert fmt_crew_sheet1(reversed_crew) == "孙军教，陈伟学"


def test_switch_serialization_round_trips() -> None:
    text = fmt_switches({"S-02": "class_level", "S-01": "all_missions_completed"})
    assert text == "S-01=all_missions_completed；S-02=class_level"


@pytest.mark.parametrize(
    ("freq_days", "gap", "expected"),
    [(3, None, 2), (7, None, 1), (14, None, 0), (7, 8, 1), (3, 3, 3)],
)
def test_minimum_required_follows_d4_formula(
    freq_days: int, gap: int | None, expected: int
) -> None:
    """v6 §3.5.3 通式：首次截止日 `max(0, freq_days − gap)`，锚点缺失按 S-12 起算。"""
    assert minimum_required(freq_days, gap_days=gap) == expected


def test_minimum_required_rejects_non_positive_freq() -> None:
    with pytest.raises(ValueError, match="freq_days"):
        minimum_required(0, gap_days=None)


# ─────────────────────────────────────────────────────────────────────
# 工作表名、顺序、表头
# ─────────────────────────────────────────────────────────────────────
def test_sheet_names_and_order_are_fixed(rendered: tuple[Path, object]) -> None:
    wb = load_workbook(rendered[0])
    assert wb.sheetnames == list(SHEET_ORDER)


def test_headers_match_template_verbatim(rendered: tuple[Path, object], tmp_path: Path) -> None:
    """产物的表头必须与 `templates/` 的模板逐字一致。"""
    build_template(tmp_path / "tpl.xlsx")
    template = read_template_headers(tmp_path / "tpl.xlsx")
    wb = load_workbook(rendered[0], data_only=True)
    for name, expected in zip(
        SHEET_ORDER[:3], (SHEET1_HEADERS, SHEET2_HEADERS, SHEET3_HEADERS), strict=True
    ):
        assert template[name] == list(expected)
        rows = [[c for c in row if c is not None] for row in wb[name].iter_rows(values_only=True)]
        headers = [row for row in rows if row == list(expected)]
        assert headers, f"{name} 里找不到逐字匹配模板的表头行"


def test_template_contract_lists_seven_blocks() -> None:
    contract = template_contract()
    assert contract["blocks"] == list(BLOCK_TITLES)
    assert len(contract["blocks"]) == 7
    assert contract["meta_labels"] == list(REQUIRED_META_LABELS)


# ─────────────────────────────────────────────────────────────────────
# 时间列必须是文本，不是 Excel 序列号（v6 §4.3 点名的断言）
# ─────────────────────────────────────────────────────────────────────
def test_time_cells_are_text_not_excel_serial(rendered: tuple[Path, object]) -> None:
    wb = load_workbook(rendered[0])
    checked = 0
    for cell in [c for row in wb[SHEET_ORDER[0]].iter_rows(min_col=1, max_col=2) for c in row]:
        if isinstance(cell.value, str) and ":" in cell.value:
            assert cell.data_type == "s", f"{cell.coordinate} 不是文本单元格"
            assert cell.number_format == "@"
            checked += 1
    assert checked >= 20, f"只检查到 {checked} 个时间单元格，样本不足"


# ─────────────────────────────────────────────────────────────────────
# Sheet 4 七区块
# ─────────────────────────────────────────────────────────────────────
def _sheet4_rows(path: Path) -> list[list[str]]:
    wb = load_workbook(path, data_only=True)
    return [
        ["" if c is None else str(c) for c in row]
        for row in wb[SHEET_ORDER[3]].iter_rows(values_only=True)
    ]


def _block_body(rows: list[list[str]], title: str, *, header_rows: int = 1) -> list[list[str]]:
    """取某个区块的数据行（不含标题行与表头行，遇到下一个区块标题即止）。"""
    heads = [r[0] for r in rows]
    start = heads.index(title) + 1
    end = next((i for i, h in enumerate(heads) if h in BLOCK_TITLES and i >= start), len(rows))
    body = [r for r in rows[start:end] if any(c.strip() for c in r)]
    return body[header_rows:]


def test_all_seven_blocks_present_and_non_empty(rendered: tuple[Path, object]) -> None:
    rows = _sheet4_rows(rendered[0])
    heads = [r[0] for r in rows]
    for title in BLOCK_TITLES:
        assert title in heads, f"缺少 {title}"
    # 每个区块标题之后至少有一行非空内容（区块1/6 无表头行，故 header_rows=0）
    for i, title in enumerate(BLOCK_TITLES):
        header_rows = 0 if i in (0, 5) else 1
        assert _block_body(rows, title, header_rows=header_rows), f"{title} 是空区块"


def test_block1_carries_every_field_needed_for_readback(rendered: tuple[Path, object]) -> None:
    rows = _sheet4_rows(rendered[0])
    labels = {r[0] for r in rows}
    for label in REQUIRED_META_LABELS:
        assert label in labels, f"区块1 缺少 {label}"
    for label in ("求解状态", "求解耗时(s)", "目标值", "gap", "worker 数", "seed", "生成时间"):
        assert label in labels, f"区块1 缺少 {label}（v6 §10.4）"


def test_block4_discloses_all_blocked_items(rendered: tuple[Path, object]) -> None:
    """披露率 100%（v6 §0.3）：阻塞项一条都不许省。"""
    path, bundle = rendered
    body = _block_body(_sheet4_rows(path), BLOCK_TITLES[3])
    assert len(body) == len(bundle.plan.blocked_items) == 7  # type: ignore[attr-defined]
    for item in bundle.plan.blocked_items:  # type: ignore[attr-defined]
        assert any(
            r[0] == f"{PERSON_NAME[item.person_id]}({item.person_id})" and r[1] == item.mission_id
            for r in body
        )


def test_block6_always_declares_the_authorized_rewrite(rendered: tuple[Path, object]) -> None:
    """S-11 开关为 on 时，「授权改写声明」是强制项（v6 §10.4 / R17）。"""
    rows = _sheet4_rows(rendered[0])
    declaration = [r for r in rows if r[0] == S11_DECLARATION_LABEL]
    assert declaration, "区块6 缺少授权改写声明行"
    assert "S-11" in declaration[0][1] and declaration[0][1] != DASH


def test_block7_has_one_row_per_sortie(rendered: tuple[Path, object]) -> None:
    path, bundle = rendered
    body = _block_body(_sheet4_rows(path), BLOCK_TITLES[6])
    assert len(body) == len(bundle.plan.sorties)  # type: ignore[attr-defined]
    assert {r[0] for r in body} == {s.sortie_id for s in bundle.plan.sorties}  # type: ignore[attr-defined]


def test_block3_does_not_charge_debt_to_blocked_combinations(rendered: tuple[Path, object]) -> None:
    """先修未满足 → 本周无要求（§3.6 BLOCKED ≠ 欠账），否则读者会以为 Tier 0 欠了账。"""
    body = _block_body(_sheet4_rows(rendered[0]), BLOCK_TITLES[2])
    blocked_rows = [r for r in body if r[2] == "先修未满足"]
    assert blocked_rows, "样本里应当有先修未满足的行"
    assert all(r[4] == "0" and r[6] == "0" for r in blocked_rows)


def test_block3_shows_dash_when_anchor_is_null(rendered: tuple[Path, object]) -> None:
    """`上次执行` 为 `—` 表示锚点 NULL（S-12），不许拿当前日期填充。"""
    body = _block_body(_sheet4_rows(rendered[0]), BLOCK_TITLES[2])
    assert body and all(r[7] == DASH for r in body)


# ─────────────────────────────────────────────────────────────────────
# 闸门3：回读反解 == 源对象
# ─────────────────────────────────────────────────────────────────────
def test_workbook_round_trips_to_the_same_plan(rendered: tuple[Path, object]) -> None:
    path, bundle = rendered
    report = verify_workbook(path, bundle.plan, ctx=bundle.ctx)  # type: ignore[attr-defined]
    assert report.diff == []
    assert report.passed


def test_runway_and_recurrent_round_trip_from_block7(rendered: tuple[Path, object]) -> None:
    """`runway_id` 与 `is_recurrent` 只在区块7 出现，必须能原样反解回来。"""
    path, bundle = rendered
    name_map = {c.name: c.person_id for s in bundle.plan.sorties for c in s.crew}  # type: ignore[attr-defined]
    parsed = parse_workbook(path, name_map=name_map)
    assert parsed.errors == []
    assert parsed.plan is not None
    source = {s.sortie_id: (s.runway_id, s.is_recurrent) for s in bundle.plan.sorties}  # type: ignore[attr-defined]
    got = {s.sortie_id: (s.runway_id, s.is_recurrent) for s in parsed.plan.sorties}
    assert got == source
    assert source[RECURRENT_SORTIE_ID] == ("RWY-1", True)
    assert deep_diff(source, got) == []


def test_relaxed_plan_round_trips_including_debts(tmp_path: Path) -> None:
    """带欠账的方案：区块3 的「松弛档」列是 `TrainingDebt` 的唯一反解来源。"""
    from tests.fixtures.report_bundle import relaxed_bundle

    bundle = relaxed_bundle()
    path = tmp_path / "relaxed.xlsx"
    render_workbook(path, bundle, readback_passed=True)
    report = verify_workbook(path, bundle.plan, ctx=bundle.ctx)
    assert report.diff == []


def test_plan_without_recurrent_sortie_round_trips(tmp_path: Path) -> None:
    bundle = sample_bundle(with_recurrent=False)
    path = tmp_path / "plain.xlsx"
    render_workbook(path, bundle, readback_passed=True)
    assert verify_workbook(path, bundle.plan, ctx=bundle.ctx).diff == []
