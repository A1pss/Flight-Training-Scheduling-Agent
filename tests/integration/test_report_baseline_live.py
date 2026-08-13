"""集成测试：基准周 2026W02 的四表产物（v6 §10 全章 + §4.3 闸门3）。

**直连裸装 PG（127.0.0.1:5433）**，读 `--baseline` 落下来的 ACTIVE 快照，
真跑一次求解 → 真跑 14 条校验 → 真渲染 xlsx → 真回读。
跑之前必须 `alembic upgrade head` 且 `python -m backend.ingestion.cli --baseline`。

这里的每一个数字都来自实际运行（铁律 6）：14 架次、7 条阻塞项、OPTIMAL，
与 v6 §1.4 的纸面推演、M2-A 的实测逐项吻合。
"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import date, datetime, timedelta, timezone
from importlib.metadata import version
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.core.db import session_scope
from backend.ingestion.loader import active_snapshot_id
from backend.nodes.compile_spec import compile_spec
from backend.report.archive import ArchiveResult, archive_plan, code_version_from_git
from backend.report.bundle import ApprovalInfo, ProvenanceInfo, ReportBundle
from backend.report.excel import render_workbook
from backend.report.manifest import load_manifest, missing_reproducibility_fields
from backend.report.naming import parse_name
from backend.report.verify import verify_existing
from backend.schemas.plan import SchedulePlan
from backend.solver.solve import solve
from backend.validator import load_context, run_all_checks, verify_format
from backend.validator.workbook import BLOCK_TITLES, SHEET_ORDER, deep_diff, parse_workbook

pytestmark = pytest.mark.integration

BASELINE_WEEK = date(2026, 1, 5)
GENERATED_AT = datetime(2026, 8, 13, 9, 0, 0, tzinfo=timezone(timedelta(hours=8)))

#: v6 §1.4.3 / M2-A 实测
EXPECTED_SORTIES = 14
EXPECTED_BLOCKED = 7

#: **刻意给 240s 而不是默认预算**（沿用 M2-D 对可复现性用例的同一处理）。
#:
#: 本模块测的是**报告层**，求解只是「弄一份真方案来渲染」。而报告层要断言的两件事
#: —— 与 M2-A 实测一致（`OPTIMAL` / 14 架次）、以及同输入可复现 —— 都以「求解跑完了」
#: 为前提：被预算截断的 `FEASIBLE` 切在哪一步取决于机器当时有多忙，v6 §3.11.1 明说
#: 它不保证一致。基准周在开发机上 18~21s 解完，**叠加 coverage 插桩后实测会顶到
#: 当时的 30s 预算之外**（2026-08-13 全量跑：`FEASIBLE` + 两次求解 `content_sha256` 不同），
#: 典型的贴着预算抛硬币。
#:
#: **这不是放宽断言**：状态仍然必须 `OPTIMAL`、架次仍然必须 14、两次求解的
#: `content_sha256` 仍然必须逐字节相同。改的只是「给求解器多少时间把最优性证完」，
#: 证完立刻返回，快的机器上一秒都不会多花。「默认预算够不够」（`Z-13` 后为 60s）由
#: `test_solver_baseline_live.py::test_baseline_week_is_optimal` 单独把关。
SOLVE_TIME_LIMIT_S = 240.0


@pytest.fixture(scope="module")
def session() -> Iterator[Session]:
    with session_scope() as s:
        yield s
        s.rollback()


@pytest.fixture(scope="module")
def snapshot(session: Session) -> str:
    snap = active_snapshot_id(session)
    assert snap, "库里没有 ACTIVE 快照 —— 先跑 `python -m backend.ingestion.cli --baseline`"
    return snap


@pytest.fixture(scope="module")
def bundle(session: Session, snapshot: str) -> ReportBundle:
    """基准周求解一次 + 校验一次，全模块共用。"""
    spec = compile_spec(
        session,
        snapshot_id=snapshot,
        week_start=BASELINE_WEEK,
        time_limit_s=SOLVE_TIME_LIMIT_S,
    )
    outcome = solve(spec)
    assert outcome.plan is not None, f"基准周应当可解，实际 {outcome.status}"
    ctx = load_context(session, snapshot_id=snapshot, week_start=BASELINE_WEEK)
    return ReportBundle(
        plan=outcome.plan,
        ctx=ctx,
        validation=run_all_checks(outcome.plan, ctx),
        stats=outcome.stats,
        generated_at=GENERATED_AT,
        format_report=verify_format(outcome.plan, ctx),
        plan_type="WEEKLY",
        plan_status="DRAFT",
        approval=ApprovalInfo(),
        provenance=ProvenanceInfo.from_settings(
            code_version=code_version_from_git(),
            # CP-SAT 版本取真实依赖（用 importlib.metadata，不 import ortools）
            solver_version=version("ortools"),
        ),
    )


@pytest.fixture(scope="module")
def archived(bundle: ReportBundle, tmp_path_factory: pytest.TempPathFactory) -> ArchiveResult:
    return archive_plan(bundle, root=tmp_path_factory.mktemp("plans"), now=GENERATED_AT)


# ─────────────────────────────────────────────────────────────────────
# 基准周本身
# ─────────────────────────────────────────────────────────────────────
def test_baseline_week_matches_the_m2a_measurement(bundle: ReportBundle) -> None:
    assert bundle.stats.status == "OPTIMAL"
    assert len(bundle.plan.sorties) == EXPECTED_SORTIES
    assert len(bundle.plan.blocked_items) == EXPECTED_BLOCKED
    assert bundle.validation.all_passed
    assert bundle.format_report is not None and bundle.format_report.passed


# ─────────────────────────────────────────────────────────────────────
# 闸门3：写出 → 回读 → 深度相等
# ─────────────────────────────────────────────────────────────────────
def test_archived_workbook_passes_gate3(archived: ArchiveResult, bundle: ReportBundle) -> None:
    report = verify_existing(archived.xlsx, bundle)
    assert report.diff == []
    assert report.passed


def test_runway_and_recurrent_round_trip(archived: ArchiveResult, bundle: ReportBundle) -> None:
    """`runway_id` / `is_recurrent` 只落在 Sheet 4 区块7，必须能原样反解。"""
    name_map = {p.name: p.person_id for p in bundle.ctx.sorted_persons()}
    parsed = parse_workbook(archived.xlsx, name_map=name_map)
    assert parsed.errors == []
    assert parsed.plan is not None
    source = {s.sortie_id: (s.runway_id, s.is_recurrent) for s in bundle.plan.sorties}
    got = {s.sortie_id: (s.runway_id, s.is_recurrent) for s in parsed.plan.sorties}
    assert deep_diff(source, got) == []
    assert set(source.values()) & {("RWY-1", False), ("RWY-2", False)}


def test_time_columns_are_text(archived: ArchiveResult) -> None:
    wb = load_workbook(archived.xlsx)
    cells = [
        c
        for row in wb[SHEET_ORDER[0]].iter_rows(min_col=1, max_col=2)
        for c in row
        if isinstance(c.value, str) and ":" in c.value
    ]
    assert len(cells) == EXPECTED_SORTIES * 2
    assert {c.data_type for c in cells} == {"s"}


def test_sheet4_has_all_seven_blocks(archived: ArchiveResult) -> None:
    wb = load_workbook(archived.xlsx, data_only=True)
    heads = [row[0] for row in wb[SHEET_ORDER[3]].iter_rows(values_only=True)]
    assert [h for h in heads if h in BLOCK_TITLES] == list(BLOCK_TITLES)


def test_block4_lists_the_seven_baseline_blocked_items(
    archived: ArchiveResult, bundle: ReportBundle
) -> None:
    wb = load_workbook(archived.xlsx, data_only=True)
    rows = [
        ["" if c is None else str(c) for c in row]
        for row in wb[SHEET_ORDER[3]].iter_rows(values_only=True)
    ]
    heads = [r[0] for r in rows]
    start = heads.index(BLOCK_TITLES[3]) + 2
    end = next(i for i, h in enumerate(heads) if h in BLOCK_TITLES and i >= start)
    body = [r for r in rows[start:end] if r[0]]
    assert len(body) == EXPECTED_BLOCKED
    assert {(r[0].split("(")[1].rstrip(")"), r[1]) for r in body} == {
        (b.person_id, b.mission_id) for b in bundle.plan.blocked_items
    }


# ─────────────────────────────────────────────────────────────────────
# 归档与 manifest
# ─────────────────────────────────────────────────────────────────────
def test_archive_layout_follows_v6_10_6(archived: ArchiveResult) -> None:
    assert archived.directory.name == "W02"
    assert archived.directory.parent.name == "2026"
    name = parse_name(archived.xlsx.name)
    assert (name.org, name.plan_type, name.iso_week) == ("NAU", "WEEKLY", "2026W02")
    assert (name.week_start, name.week_end) == ("20260105", "20260111")
    for path in archived.all_paths():
        assert path.exists() and path.stat().st_size > 0


def test_manifest_carries_every_reproducibility_field(archived: ArchiveResult) -> None:
    manifest = load_manifest(archived.manifest)
    assert missing_reproducibility_fields(manifest) == ()
    assert len(manifest["semantics_switches"]) == 13  # S-01~S-13
    assert manifest["solver"]["num_search_workers"] >= 1
    assert manifest["solver"]["seed"] == 42


def test_second_export_gets_a_new_version(bundle: ReportBundle, archived: ArchiveResult) -> None:
    root = archived.directory.parent.parent
    again = archive_plan(bundle, root=root, now=GENERATED_AT)
    assert again.name.version == archived.name.version + 1
    assert again.xlsx != archived.xlsx
    assert archived.xlsx.exists(), "旧版本不得被覆盖"


# ─────────────────────────────────────────────────────────────────────
# 可复现性（铁律 9）
# ─────────────────────────────────────────────────────────────────────
def test_same_inputs_reproduce_the_same_plan(
    session: Session, snapshot: str, bundle: ReportBundle
) -> None:
    """按 manifest 里的 snapshot / ruleset / semantics / seed 再解一次，方案逐字节相同。"""
    spec = compile_spec(
        session,
        snapshot_id=snapshot,
        week_start=BASELINE_WEEK,
        time_limit_s=SOLVE_TIME_LIMIT_S,
        materialize=False,  # 只读复跑，不重写 training_progress
    )
    outcome = solve(spec)
    assert outcome.plan is not None
    assert outcome.status == "OPTIMAL", f"可复现性以「求解跑完」为前提，实际 {outcome.status}"
    assert outcome.plan.content_sha256 == bundle.plan.content_sha256
    assert _canonical(outcome.plan) == _canonical(bundle.plan)


def test_rendering_is_deterministic(bundle: ReportBundle, tmp_path: Path) -> None:
    """同一份 bundle 渲染两次，四张表逐格相同（xlsx 的 zip 时间戳不参与比较）。"""
    first, second = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    render_workbook(first, bundle, readback_passed=True)
    render_workbook(second, bundle, readback_passed=True)
    assert _cells(first) == _cells(second)


def _canonical(plan: SchedulePlan) -> str:
    import json

    return json.dumps(plan.model_dump(mode="json"), ensure_ascii=False, sort_keys=True)


def _cells(path: Path) -> dict[str, list[list[object]]]:
    wb = load_workbook(path, data_only=True)
    return {
        name: [list(row) for row in wb[name].iter_rows(values_only=True)] for name in SHEET_ORDER
    }
