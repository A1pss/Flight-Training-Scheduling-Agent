"""按 `backend.validator.workbook` 的回读契约手工写出一份 xlsx。

**这不是 M3 的 Excel 写出模块**（那一版还要负责字体、底纹、列宽、合并单元格与
Sheet 4 的展示型区块）。这里只写「回读所需的那一份」，用途是把闸门3 的比对逻辑
在本窗口测通：写出 → 回读 → 深度相等。M3 落地时直接 import
`backend.validator.workbook` 的常量，并让 `verify_workbook` 绿灯即可。

时间列一律写成 `HH:MM` **文本**（v6 §4.3 的点名断言），日期列写 ISO 文本。
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from backend.schemas.plan import SchedulePlan, Sortie
from backend.validator.schema import WEEKDAY_ORDER
from backend.validator.workbook import (
    BLOCK2_HEADERS,
    BLOCK3_HEADERS,
    BLOCK4_HEADERS,
    BLOCK7_HEADERS,
    BLOCK_TITLES,
    META_COVER,
    META_ISO_WEEK,
    META_PLAN_ID,
    META_RULESET,
    META_RUNWAY_MODEL,
    META_SEMANTICS,
    META_SHA,
    META_SNAPSHOT,
    META_SWITCHES,
    META_TIER,
    NOT_RECURRENT_MARK,
    RECURRENT_MARK,
    ROLE_SUFFIX,
    SHEET1_HEADERS,
    SHEET2_HEADERS,
    SHEET3_HEADERS,
    SHEET_ORDER,
    SWITCH_KV,
    SWITCH_SEP,
)

AIRSPACE_NAMES: Mapping[str, str] = {
    "SAA": "Small Area A",
    "SAB": "Small Area B",
    "IFR": "IFR Route",
    "RT1": "Route 1",
    "RT2": "Route 2",
    "RNG": "Range Area",
}


def _hhmm(value: Any) -> str:
    return f"{value.hour:02d}:{value.minute:02d}"


def _mission_cell(s: Sortie, *, with_airspace: bool) -> str:
    base = f"{s.mission_name} ({s.mission_id})"
    if with_airspace:
        return f"{base}（{AIRSPACE_NAMES.get(s.airspace_id, s.airspace_id)}）"
    return base


def _sheet1_crew(s: Sortie) -> str:
    return "，".join(f"{c.name}{ROLE_SUFFIX[c.role]}" for c in s.crew)


def _sorted(sorties: Sequence[Sortie]) -> list[Sortie]:
    return sorted(sorties, key=lambda s: (s.date, s.takeoff, s.sortie_id))


def write_workbook(
    plan: SchedulePlan,
    path: Path,
    *,
    aircraft_types: Mapping[str, str],
    sheet_names: Sequence[str] | None = None,
    time_as_native: bool = False,
) -> Path:
    """写出一份符合回读契约的 xlsx。

    `sheet_names` / `time_as_native` 只为测试「版式被写坏」的分支存在：
    前者改工作表名，后者把时间列写成 `datetime.time`（Excel 序列号形态）。
    """
    wb = Workbook()
    names = tuple(sheet_names or SHEET_ORDER)
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = names[0]
    ws2 = wb.create_sheet(names[1])
    ws3 = wb.create_sheet(names[2])
    ws4 = wb.create_sheet(names[3])
    ordered = _sorted(plan.sorties)

    def _t(value: Any) -> Any:
        return value if time_as_native else _hhmm(value)

    # ── Sheet 1 · 分日飞行计划表（按星期分组，组内按起飞升序）────────────
    for weekday in WEEKDAY_ORDER:
        rows = [s for s in ordered if s.weekday == weekday]
        if not rows:
            continue
        ws1.append([weekday])
        ws1.append(list(SHEET1_HEADERS))
        for s in rows:
            ws1.append(
                [
                    _t(s.takeoff),
                    _t(s.landing),
                    s.aircraft_id,
                    _mission_cell(s, with_airspace=True),
                    _sheet1_crew(s),
                ]
            )
        ws1.append([])

    # ── Sheet 2 · 飞行员训练时间表（人员 → 星期）────────────────────────
    by_person: dict[str, list[Sortie]] = {}
    names_by_person: dict[str, str] = {}
    for s in ordered:
        for c in s.crew:
            by_person.setdefault(c.person_id, []).append(s)
            names_by_person[c.person_id] = c.name
    for person_id in sorted(by_person):
        ws2.append([names_by_person[person_id]])
        ws2.append(list(SHEET2_HEADERS))
        for weekday in WEEKDAY_ORDER:
            rows = [s for s in by_person[person_id] if s.weekday == weekday]
            if not rows:
                continue
            ws2.append([None, weekday])
            for s in rows:
                role = next(c.role for c in s.crew if c.person_id == person_id)
                ws2.append(
                    [
                        None,
                        None,
                        f"{_hhmm(s.takeoff)}-{_hhmm(s.landing)}",
                        _mission_cell(s, with_airspace=False),
                        f"({s.aircraft_id}/{role})",
                    ]
                )
        ws2.append([])

    # ── Sheet 3 · 飞机排班表（机号 → 星期）──────────────────────────────
    by_aircraft: dict[str, list[Sortie]] = {}
    for s in ordered:
        by_aircraft.setdefault(s.aircraft_id, []).append(s)
    for ac_id in sorted(by_aircraft):
        ws3.append([ac_id])
        ws3.append(list(SHEET3_HEADERS))
        for weekday in WEEKDAY_ORDER:
            rows = [s for s in by_aircraft[ac_id] if s.weekday == weekday]
            if not rows:
                continue
            ws3.append([None, weekday])
            for s in rows:
                # 双人按 (教员, 学员) 顺序；单人只写姓名，角色由复训标记还原
                members = sorted(s.crew, key=lambda c: 0 if c.role == "教员" else 1)
                ws3.append(
                    [
                        None,
                        None,
                        _t(s.takeoff),
                        _mission_cell(s, with_airspace=False),
                        "(" + "/".join(c.name for c in members) + ")",
                    ]
                )
        ws3.append([])

    # ── Sheet 4 · 合规与解释报告（七区块）───────────────────────────────
    switches = SWITCH_SEP.join(
        f"{k}{SWITCH_KV}{v}" for k, v in sorted(plan.semantics_switches.items())
    )
    ws4.append([BLOCK_TITLES[0]])
    for label, value in (
        (META_PLAN_ID, plan.plan_id),
        (META_ISO_WEEK, plan.iso_week),
        (META_COVER, f"{plan.week_start.isoformat()} ~ {plan.week_end.isoformat()}"),
        (META_SNAPSHOT, plan.snapshot_id),
        (META_RULESET, plan.ruleset_version),
        (META_SEMANTICS, plan.semantics_version),
        (META_SWITCHES, switches),
        (META_RUNWAY_MODEL, f"{plan.runway_model}（RWY-1: JL-8/JL-9；RWY-2: JL-8）"),
        (META_TIER, f"Tier {plan.relaxation_tier}（全硬约束）"),
        (META_SHA, plan.content_sha256),
    ):
        ws4.append([label, value])
    ws4.append([])

    ws4.append([BLOCK_TITLES[1]])
    ws4.append(list(BLOCK2_HEADERS))
    ws4.append([])

    ws4.append([BLOCK_TITLES[2]])
    ws4.append(list(BLOCK3_HEADERS))
    for d in sorted(plan.debts, key=lambda d: (d.person_id, d.mission_id)):
        ws4.append(
            [
                f"{names_by_person.get(d.person_id, d.person_id)}({d.person_id})",
                d.mission_id,
                "未完成",
                "—",
                d.required,
                d.scheduled,
                d.debt,
                "—",
                "—",
                d.relaxed_by,
            ]
        )
    ws4.append([])

    ws4.append([BLOCK_TITLES[3]])
    ws4.append(list(BLOCK4_HEADERS))
    for b in sorted(plan.blocked_items, key=lambda b: (b.person_id, b.mission_id)):
        ws4.append(
            [
                f"{names_by_person.get(b.person_id, b.person_id)}({b.person_id})",
                b.mission_id,
                b.reason,
                "、".join(b.missing_prereqs) or "—",
                "完成先修后",
            ]
        )
    ws4.append([])

    ws4.append([BLOCK_TITLES[4]])
    ws4.append(["对象", "架次", "飞行时长(分)", "利用率", "备注"])
    ws4.append([])

    ws4.append([BLOCK_TITLES[5]])
    ws4.append(
        ["使用的松弛", "无" if plan.relaxation_tier == 0 else f"Tier {plan.relaxation_tier}"]
    )
    ws4.append(
        [
            "授权改写声明",
            "S-11：成熟飞行员到期资质转复训，系对 rules.pdf 约束2 字面语义的业务方授权改写"
            "（2026-08-06 裁定），非校验器漏判",
        ]
    )
    ws4.append([])

    ws4.append([BLOCK_TITLES[6]])
    ws4.append(list(BLOCK7_HEADERS))
    for s in ordered:
        ws4.append(
            [
                s.sortie_id,
                s.date.isoformat(),
                _t(s.takeoff),
                s.aircraft_id,
                aircraft_types.get(s.aircraft_id, ""),
                s.runway_id,
                s.airspace_id,
                RECURRENT_MARK if s.is_recurrent else NOT_RECURRENT_MARK,
            ]
        )

    wb.save(path)
    return path


__all__ = ["AIRSPACE_NAMES", "write_workbook"]
